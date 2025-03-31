import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
from openpyxl import Workbook
import pandas as pd
import csv
import os
import re

class ImportadorExtratos:
    def __init__(self, root):
        self.root = root
        self.root.title("Importador de Extratos Bancários")
        self.root.geometry("1000x600")
        self.root.config(bg='#D9D9D9')
        self.root.state("zoomed")
        self.root.iconbitmap(r"C:\Users\regina.santos\Desktop\Automação\Judite\icon.ico")
        self.root.protocol("WM_DELETE_WINDOW", self.fechar_janela)

        for i in range(6):
            root.grid_columnconfigure(i, weight=0)
        for i in range(6):
            root.grid_rowconfigure(i, weight=0)

        self.janelas_filhas = []

        self.title_label = tk.Label(root, text="Importador de Extratos", font=("Roboto", 17, "bold"), bg='#D9D9D9')
        self.title_label.grid(row=3, column=0, padx=10, sticky="w")

        self.btn_importar = tk.Button(root, text="Importar", command=self.mostrar_selecao_conta, font=("Roboto", 10), bg="#4CAF50", fg="white", relief="groove", padx=5, pady=3)
        self.btn_importar.grid(row=4, column=0, padx=10, sticky="w")

        self.btn_limpar = tk.Button(root, text="Limpar Tudo", command=self.confirmar_limpar_dados, font=("Roboto", 10), bg="#FF5733", fg="white", relief="groove", padx=5, pady=3)
        self.btn_limpar.grid(row=4, column=0, padx=80, sticky="w")

        self.btn_classificar1p = tk.Button(root, text="Classificar 1P", command=self.classificar_dados, font=("Roboto", 10), bg="#4CAF50", fg="white", relief="groove", padx=5, pady=3)
        self.btn_classificar1p.grid(row=2, column=3, padx=90, sticky="e")

        self.btn_limpar1p = tk.Button(root, text="Limpar 1P", command=self.limpar_1p, font=("Roboto", 10), bg="#4CAF50", fg="white", relief="groove", padx=5, pady=3)
        self.btn_limpar1p.grid(row=2, column=3, padx=5, sticky="e")

        self.btn_exportar = tk.Button(root, text="Exportar", command=self.exportar_dados, font=("Roboto", 10), bg="#4CAF50", fg="white", relief="groove", padx=5, pady=3)
        self.btn_exportar.grid(row=2, column=4, sticky="w")

        self.saldo_inicial_label = tk.Label(root, text="Saldo Inicial Importado>", font=("Roboto", 11), bg='#D9D9D9', anchor='e')
        self.saldo_inicial_label.grid(row=2, column=1, padx=1, sticky="w")
        self.saldo_inicial_entry = tk.Entry(root, font=("Roboto", 10), bd=1, relief="solid", width=17)
        self.saldo_inicial_entry.grid(row=2, column=1, columnspan=2, padx=160, sticky="w")

        self.saldo_final_label = tk.Label(root, text="Saldo Final Importado>", font=("Roboto", 11), bg='#D9D9D9', anchor='e')
        self.saldo_final_label.grid(row=3, column=1, padx=3, sticky="w")
        self.saldo_final_entry = tk.Entry(root, font=("Roboto", 10), bd=1, relief="solid", width=17)
        self.saldo_final_entry.grid(row=3, column=1, columnspan=2, padx=160, sticky="w")

        self.saldo_final_calculado_label = tk.Label(root, text="Saldo Final Calculado>", font=("Roboto", 11), bg='#D9D9D9', anchor='e')
        self.saldo_final_calculado_label.grid(row=4, column=1, padx=3, sticky="w")
        self.saldo_final_calculado_entry = tk.Entry(root, font=("Roboto", 10), bd=1, relief="solid", width=17)
        self.saldo_final_calculado_entry.grid(row=4, column=1, columnspan=2, padx=160, sticky="w")

        self.diferenca_label = tk.Label(root, text="Diferença>", font=("Roboto", 11), bg='#D9D9D9', anchor='e')
        self.diferenca_label.grid(row=5, column=1, padx=82, sticky="w")
        self.diferenca_entry = tk.Entry(root, font=("Roboto", 10), bd=1, relief="solid", width=17)
        self.diferenca_entry.grid(row=5, column=1, columnspan=2, padx=160, sticky="w")

        self.empresa_label = tk.Label(root, text="Empresa>", font=("Roboto", 11), bg='#D9D9D9', anchor='w')
        self.empresa_label.grid(row=3, column=2, padx=51, sticky="w")
        self.empresa_entry = tk.Entry(root, font=("Roboto", 10), bd=1, relief="solid", width=10)
        self.empresa_entry.grid(row=3, column=2, padx=125, sticky="w")

        self.conta_label = tk.Label(root, text="Conta>", font=("Roboto", 11), bg='#D9D9D9', anchor='w')
        self.conta_label.grid(row=4, column=2, padx=72, sticky="w")
        self.conta_entry = tk.Entry(root, font=("Roboto", 10), bd=1, relief="solid", width=10)
        self.conta_entry.grid(row=4, column=2, padx=125, sticky="w")

        self.centro_custo_label = tk.Label(root, text="C/Custo>", font=("Roboto", 11), bg='#D9D9D9', anchor='w')
        self.centro_custo_label.grid(row=5, column=2, padx=54, sticky="w")
        self.centro_custo_entry = tk.Entry(root, font=("Roboto", 10), bd=1, relief="solid", width=10)
        self.centro_custo_entry.grid(row=5, column=2, padx=125, sticky="w")

        self.banco_label = tk.Label(root, text="Banco>", font=("Roboto", 11), bg='#D9D9D9', anchor='w')
        self.banco_label.grid(row=4, column=2, padx=1, sticky="e")
        self.banco_entry = tk.Entry(root, font=("Roboto", 10), bd=1, relief="solid", width=20)
        self.banco_entry.grid(row=4, column=3, padx=1, sticky="w")

        self.agencia_conta_label = tk.Label(root, text="Agência/Conta>", font=("Roboto", 11), bg='#D9D9D9', anchor='w')
        self.agencia_conta_label.grid(row=5, column=2, padx=1, sticky="e")
        self.agencia_conta_entry = tk.Entry(root, font=("Roboto", 10), bd=1, relief="solid", width=20)
        self.agencia_conta_entry.grid(row=5, column=3, padx=1, sticky="w")

        self.saldo_final_contabil_label = tk.Label(root, text="Saldo Final Contábil>", font=("Roboto", 11), bg='#D9D9D9', anchor='w')
        self.saldo_final_contabil_label.grid(row=4, column=3, columnspan=4, padx=95, sticky="e")
        self.saldo_final_contabil_entry = tk.Entry(root, font=("Roboto", 10), bd=1, relief="solid", width=10)
        self.saldo_final_contabil_entry.grid(row=4, column=3, columnspan=4, padx=10, sticky="e")

        self.diferenca_extrato_bancario_label = tk.Label(root, text="Diferença c/Extrato Bancário>", font=("Roboto", 11), bg='#D9D9D9', anchor='w')
        self.diferenca_extrato_bancario_label.grid(row=5, column=3, columnspan=4, padx=95, sticky="e")
        self.diferenca_extrato_bancario_entry = tk.Entry(root, font=("Roboto", 10), bd=1, relief="solid", width=10)
        self.diferenca_extrato_bancario_entry.grid(row=5, column=3, columnspan=4, padx=10, sticky="e")

        self.linhadeajuda_label = tk.Label(root, text="Coluna 0", fg='#D9D9D9', font=("Roboto", 10), bg='#D9D9D9')
        self.linhadeajuda_label.grid(row=1, column=0)
        self.linhadeajuda_label = tk.Label(root, text="Coluna 1", fg='#D9D9D9', font=("Roboto", 10), bg='#D9D9D9')
        self.linhadeajuda_label.grid(row=1, column=1)
        self.linhadeajuda_label = tk.Label(root, text="Coluna 2", fg='#D9D9D9', font=("Roboto", 10), bg='#D9D9D9')
        self.linhadeajuda_label.grid(row=1, column=2)
        self.linhadeajuda_label = tk.Label(root, text="Coluna 3", fg='#D9D9D9', font=("Roboto", 10), bg='#D9D9D9')
        self.linhadeajuda_label.grid(row=1, column=3)
        self.linhadeajuda_label = tk.Label(root, text="Coluna 4", fg='#D9D9D9', font=("Roboto", 10), bg='#D9D9D9')
        self.linhadeajuda_label.grid(row=1, column=4)
        self.linhadeajuda_label = tk.Label(root, text="Coluna 5", fg='#D9D9D9', font=("Roboto", 10), bg='#D9D9D9')
        self.linhadeajuda_label.grid(row=1, column=5)

        self.tree = ttk.Treeview(root, columns=(
            "DataLEB", "DescriçãoLEB", "NDocLEB", "ValorLEB", "SaldoLEB",
            "LançamentoLC", "DataLC", "DébitoLC", "D-C/CLC", "CréditoLC",
            "C-C/CLC", "CNPJLC", "HistóricoLC", "ValorLC"
        ), show="headings")

        self.tree.heading("DataLEB", text="Data")
        self.tree.heading("DescriçãoLEB", text="Descrição")
        self.tree.heading("NDocLEB", text="N° Doc")
        self.tree.heading("ValorLEB", text="Valor")
        self.tree.heading("SaldoLEB", text="Saldo")
        self.tree.heading("LançamentoLC", text="Lançamento")
        self.tree.heading("DataLC", text="Data")
        self.tree.heading("DébitoLC", text="Débito")
        self.tree.heading("D-C/CLC", text="D-C/C")
        self.tree.heading("CréditoLC", text="Crédito")
        self.tree.heading("C-C/CLC", text="C-C/C")
        self.tree.heading("CNPJLC", text="CNPJ")
        self.tree.heading("HistóricoLC", text="Histórico")
        self.tree.heading("ValorLC", text="Valor")

        self.tree.column("DataLEB", width=60)
        self.tree.column("DescriçãoLEB", width=150)
        self.tree.column("NDocLEB", width=60)
        self.tree.column("ValorLEB", width=60)
        self.tree.column("SaldoLEB", width=60)
        self.tree.column("LançamentoLC", width=100)
        self.tree.column("DataLC", width=60)
        self.tree.column("DébitoLC", width=80)
        self.tree.column("D-C/CLC", width=80)
        self.tree.column("CréditoLC", width=80)
        self.tree.column("C-C/CLC", width=80)
        self.tree.column("CNPJLC", width=80)
        self.tree.column("HistóricoLC", width=100)
        self.tree.column("ValorLC", width=80)
        self.tree.grid(row=7, column=0, columnspan=6, pady=5, padx=10, sticky="nsew")

        root.grid_rowconfigure(7, weight=1)
        root.grid_columnconfigure(0, weight=1)
        root.grid_columnconfigure(1, weight=1)
        root.grid_columnconfigure(2, weight=1)
        root.grid_columnconfigure(3, weight=1)
        root.grid_columnconfigure(4, weight=1)
        root.grid_columnconfigure(5, weight=1)

    def mostrar_selecao_conta(self):
        from telas.tela_selecao_conta import TelaSelecaoConta
        janela_selecao_conta = tk.Toplevel(self.root)
        app = TelaSelecaoConta(janela_selecao_conta, self.abrir_explorador_arquivos)
        self.janelas_filhas.append(janela_selecao_conta)

    def abrir_tela_nova_empresa(self):
        from telas.tela_nova_empresa import TelaNovaEmpresa
        if not hasattr(self, 'tela_nova_empresa') or not self.tela_nova_empresa.winfo_exists():
            self.abrir_tela_nova_empresa = tk.Toplevel(self.root)
            app = TelaNovaEmpresa(self.abrir_tela_nova_empresa, self.salvar_nova_empresa)
            self.janelas_filhas.append(self.tela_nova_empresa)

    def abrir_tela_nova_conta(self):
        from telas.tela_nova_conta import TelaNovaConta
        if not hasattr(self, 'tela_nova_conta') or not self.tela_nova_conta.winfo_exists():
            self.tela_nova_conta = tk.Toplevel(self.root)
            app = TelaNovaConta(self.tela_nova_conta, self.salvar_nova_conta)
            self.janelas_filhas.append(self.tela_nova_conta)

    def fechar_janela(self):
        for janela in self.janelas_filhas:
            if janela.winfo_exists():
                janela.destroy()
        self.root.destroy()

    def abrir_explorador_arquivos(self, empresa, conta, arquivo):
        self.empresa_entry.delete(0, tk.END)
        self.empresa_entry.insert(0, empresa.split(" - ")[0])
        self.conta_entry.delete(0, tk.END)
        self.conta_entry.insert(0, conta.split(" - ")[0])
        self.banco_entry.delete(0, tk.END)
        self.banco_entry.insert(0, conta.split(" - ")[1])

        agencia = conta.split(" - ")[3]
        conta_bancaria = conta.split(" - ")[4]
        self.agencia_conta_entry.delete(0, tk.END)
        self.agencia_conta_entry.insert(0, f"{agencia}/{conta_bancaria}")

        self.importar_dados_arquivo(arquivo)


    




    def importar_dados_arquivo(self, arquivo):
        wb = openpyxl.load_workbook(arquivo, data_only=True)
        sheet = wb.active

        saldo_inicial = sheet["F10"].value
        saldo_final = sheet["F212"].value
        saldo_final_calculado = sheet["F212"].value

        self.saldo_inicial_entry.delete(0, tk.END)
        self.saldo_inicial_entry.insert(0, saldo_inicial)

        self.saldo_final_entry.delete(0, tk.END)
        self.saldo_final_entry.insert(0, saldo_final)

        self.saldo_final_calculado_entry.delete(0, tk.END)
        self.saldo_final_calculado_entry.insert(0, saldo_final_calculado)

        dados_importados = []

        for row in range(11, sheet.max_row + 1):
            coluna_a = sheet.cell(row=row, column=1).value
            if coluna_a and isinstance(coluna_a, str) and coluna_a.strip().lower() == "total":
                break

            coluna_a = sheet.cell(row=row, column=1).value
            coluna_b = sheet.cell(row=row, column=2).value
            coluna_c = sheet.cell(row=row, column=3).value
            coluna_d = sheet.cell(row=row, column=4).value
            coluna_e = sheet.cell(row=row, column=5).value
            coluna_f = sheet.cell(row=row, column=6).value

            def tratar_valor(valor):
                if valor is None or valor == "":
                    return 0
                valor = str(valor).replace(".", "").replace(",", ".")
                try:
                    return float(valor)
                except ValueError:
                    return 0

            coluna_d = tratar_valor(coluna_d)
            coluna_e = tratar_valor(coluna_e)

            valor_total = coluna_d + coluna_e
            dados_importados.append([
                coluna_a, coluna_b, coluna_c, valor_total, coluna_f,
                "", "", "", "", "", "", "", "", ""
            ])

        df = pd.DataFrame(dados_importados, columns=[
            "DataLEB", "DescriçãoLEB", "NDocLEB", "ValorLEB", "SaldoLEB",
            "LançamentoLC", "DataLC", "DébitoLC", "D-C/CLC", "CréditoLC",
            "C-C/CLC", "CNPJLC", "HistóricoLC", "ValorLC"
        ])

        for i in self.tree.get_children():
            self.tree.delete(i)

        for _, row in df.iterrows():
            self.tree.insert("", "end", values=list(row))

    def confirmar_limpar_dados(self):
        resposta = messagebox.askyesno("Atenção", "Tem certeza que deseja limpar todos os dados?")
        if resposta:
            self.limpar_dados()

    def limpar_dados(self):
        self.saldo_inicial_entry.delete(0, tk.END)
        self.saldo_final_entry.delete(0, tk.END)
        self.saldo_final_calculado_entry.delete(0, tk.END)
        self.empresa_entry.delete(0, tk.END)
        self.conta_entry.delete(0, tk.END)
        self.diferenca_entry.delete(0, tk.END)
        self.banco_entry.delete(0, tk.END)
        self.agencia_conta_entry.delete(0, tk.END)
        self.saldo_final_contabil_entry.delete(0, tk.END)
        self.diferenca_extrato_bancario_entry.delete(0, tk.END)
        for i in self.tree.get_children():
            self.tree.delete(i)

    def exportar_dados(self):
        wb = Workbook()
        ws = wb.active

        colunas_exportar = ["LançamentoLC", "DataLC", "DébitoLC", "D-C/CLC", "CréditoLC",
                      "C-C/CLC", "CNPJLC", "HistóricoLC", "ValorLC"]

        dados = []
        for item in self.tree.get_children():
            values = self.tree.item(item, 'values')
            dados.append([values[self.tree["columns"].index(col)] for col in colunas_exportar])

        df = pd.DataFrame(dados, columns=colunas_exportar)

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Arquivos Excel", "*.xlsx")])
        if file_path:
            df.to_excel(file_path, index=False)
            messagebox.showinfo("Exportar", "Dados exportados com sucesso!")

    def classificar_dados(self):
        messagebox.showinfo("Classificar 1P", "Essa função não está disponível no momento!")

    def limpar_1p(self):
        messagebox.showinfo("Limpar 1P", "Essa função não está disponível no momento!")
