import pandas as pd

from openpyxl import load_workbook

wb = load_workbook(r"K:\013 - Integracao\2025\Agenda Controle INT\Controle de INT - 01.2025.xlsx", data_only=True)
ws = wb["GPS"]

valores_filtrados = []

for row in ws.iter_rows(min_row=2, values_only=True):
    if row[12] == "pendente":
        valores_filtrados.append(row[0])

print(valores_filtrados)









#df = pd.read_excel(r"K:\013 - Integracao\2025\Agenda Controle INT\Controle de INT - 01.2025.xlsx", sheet_name="GPS", engine="openpyxl")
#df_filtrado = df[df.iloc[:, 12] == "Pendente"]
#print(df_filtrado.iloc[:, 0])





#wb = load_workbook(r"K:\013 - Integracao\2025\Agenda Controle INT\Controle de INT - 01.2025.xlsx")
#ws = wb["GPS"]

#for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
#    print(row[0])







#df = pd.read_excel(r"K:\013 - Integracao\2025\Agenda Controle INT\Controle de INT - 01.2025.xlsx", sheet_name="GPS", engine="openpyxl")

#df_filtrado = df[df["Integração - GPS"] == "Pendente"]

#primeira_coluna = df.iloc[:, 0]


#print(primeira_coluna)