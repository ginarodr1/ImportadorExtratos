import sys
import os
import getpass
import logging

log_dir = r"K:\013 - Integracao\Macro\Importador de extratos\Logging"
os.makedirs(log_dir, exist_ok=True)
user = getpass.getuser()
log_path = os.path.join(log_dir, f"{user}.log")

logging.basicConfig(
    filename=log_path,
    filemode='a',
    format='%(asctime)s - %(levelname)s - %(message)s',
    level=logging.INFO
)

logging.info(f"Aplicação iniciada pelo usuário {user}")

import tkinter as tk
from tkinter import filedialog, messagebox, ttk, simpledialog
import openpyxl
from openpyxl import Workbook
import pandas as pd
import csv
import os
import re
import locale
from datetime import datetime, date
import traceback
import xlrd
from unidecode import unidecode
from fuzzywuzzy import process, fuzz
import bcrypt

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

def caminho_dos_csv(nome_arquivo):
    #"""Retorna o caminho absoluto do CSV com base na pasta do script"""
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), nome_arquivo)

locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

def open_main_window():
    root = tk.Tk()
    app = ImportadorExtratos(root)
    root.mainloop()

class SenhaLogin:
    def __init__(self, root, callback):
        self.root = root
        self.callback = callback
        self.root.title("Login")
        self.root.geometry("247x100")
        self.root.config(bg='#f4f4f4')
        self.root.iconbitmap(r"K:\013 - Integracao\Macro\Importador de extratos\icon.ico")

        self.titulosenha = tk.Label(root, text="Importador de Extratos está protegido.", font=("Roboto", 10), bg='#f4f4f4')
        self.titulosenha.grid(row=2, column=0, columnspan=2, pady=5, padx=10, sticky="w")

        self.senha_label = tk.Label(root, text="Senha:", font=("Roboto", 10), bg='#f4f4f4')
        self.senha_label.grid(row=3, column=0, columnspan=2, pady=5, padx=10, sticky="w")

        self.senha_entry = tk.Entry(root, show="*", font=("Roboto", 10), width=20)
        self.senha_entry.grid(row=3, column=0, padx=60, sticky="w")

        self.login_button = tk.Button(root, text="Entrar", command=self.verificar_senha, font=("Roboto", 10), bg="#4CAF50", fg="white")
        self.login_button.grid(row=4, column=0, columnspan=1, padx=27, sticky="e")

        self.senha_entry.bind("<Return>", lambda event: self.verificar_senha())

        self.root.after(0, lambda: self.root.focus_force())
        self.senha_entry.focus_set()

        self.senha_hash = b'$2b$12$jvCTLhlKP5KOXeVGkVC0PuptYaivVLDtGq34mZuEx38uMO18oGOjW' #! senha em hash gerada e armazenada

    def verificar_senha(self):
        senha = self.senha_entry.get()
        if bcrypt.checkpw(senha.encode(), self.senha_hash):
            self.root.destroy()
            self.callback()
        else:
            messagebox.showerror("Erro", "Senha incorreta!")
            self.senha_entry.delete(0, tk.END)
        
class ImportadorExtratos:
    def __init__(self, root):
        logging.info("Janela principal iniciada.")
        self.root = root
        self.root.title("Importador de Extratos Bancários")
        self.root.geometry("1000x600")
        self.root.config(bg='#313131')
        self.root.state("zoomed")
        self.root.iconbitmap(r"K:\013 - Integracao\Macro\Importador de extratos")
        self.root.protocol("WM_DELETE_WINDOW", self.fechar_janela)

        self.df_banco_dados = pd.read_csv(resource_path("lancamentoscontas1.csv"), delimiter=';')
        self.empresas_df = pd.read_csv(resource_path("empresas.csv"))
        self.contas_df = pd.read_csv(resource_path("contas.csv"))

        self.root.tk.call("source", r"K:\013 - Integracao\Macro\Forest-ttk-theme\forest-dark.tcl")
        style = ttk.Style(self.root)
        style.theme_use("forest-dark")

        for i in range(6):
            root.grid_columnconfigure(i, weight=0)
        for i in range(6):
            root.grid_rowconfigure(i, weight=0)

        self.janelas_filhas = [] #? armazenar todas as janelas filhas

        #! -------------------- TÍTULO PRINCIPAL -------------------- #

        self.title_label = tk.Label(root, text="Importador de Extratos", font=("Roboto", 17, "bold"), bg='#313131', fg="white")
        self.title_label.grid(row=3, column=0, padx=10, sticky="w")

        #! -------------------- BOTÕES PRINCIPAIS -------------------- #

        self.btn_importar = ttk.Button(root, text="Importar", command=self.mostrar_selecao_conta, width=7)
        self.btn_importar.grid(row=4, column=0, padx=10, sticky="w")

        self.btn_limpar = ttk.Button(root, text="Limpar Tudo", command=self.confirmar_limpar_dados, width=10)
        self.btn_limpar.grid(row=4, column=0, padx=87, sticky="w")

        self.btn_classificar1p = ttk.Button(root, text="Classificar 1P", command=self.classificar_dados, width=12)
        self.btn_classificar1p.grid(row=2, column=3, padx=91, sticky="e")

        self.btn_limpar1p = ttk.Button(root, text="Limpar 1P", command=self.limpar_1p, width=9)
        self.btn_limpar1p.grid(row=2, column=3, padx=2, sticky="e")

        self.btn_exportar = ttk.Button(root, text="Exportar", command=self.exportar_dados, width=7)
        self.btn_exportar.grid(row=2, column=4, sticky="w")

        #! -------------------- CAMPOS DE SALDO -------------------- #
        #* -------------- LANÇAMENTO EXTRATO BANCÁRIO -------------- #

        self.saldo_inicial_label = tk.Label(root, text="Saldo Inicial Importado", font=("Roboto", 11), bg='#313131', fg="white", anchor='e')
        self.saldo_inicial_label.grid(row=2, column=1, padx=1, sticky="w")
        self.saldo_inicial_entry = ttk.Entry(root, width=17)
        self.saldo_inicial_entry.grid(row=2, column=1, columnspan=2, padx=163, sticky="w")

        self.saldo_final_label = tk.Label(root, text="Saldo Final Importado", font=("Roboto", 11), bg='#313131', fg="white", anchor='e')
        self.saldo_final_label.grid(row=3, column=1, padx=8, sticky="w")
        self.saldo_final_entry = ttk.Entry(root, width=17)
        self.saldo_final_entry.grid(row=3, column=1, columnspan=2, padx=163, sticky="w")

        self.saldo_final_calculado_label = tk.Label(root, text="Saldo Final Calculado", font=("Roboto", 11), bg='#313131', fg="white", anchor='e')
        self.saldo_final_calculado_label.grid(row=4, column=1, padx=10, sticky="w")
        self.saldo_final_calculado_entry = ttk.Entry(root, width=17)
        self.saldo_final_calculado_entry.grid(row=4, column=1, columnspan=2, padx=163, sticky="w")

        self.diferenca_label = tk.Label(root, text="Diferença", font=("Roboto", 11), bg='#313131', fg="white", anchor='e')
        self.diferenca_label.grid(row=5, column=1, padx=89, sticky="w")
        self.diferenca_entry = ttk.Entry(root, width=17)
        self.diferenca_entry.grid(row=5, column=1, columnspan=2, padx=163, sticky="w")

        self.empresa_label = tk.Label(root, text="Empresa", font=("Roboto", 11), bg='#313131', fg="white", anchor='w')
        self.empresa_label.grid(row=3, column=2, padx=59, sticky="w")
        self.empresa_entry = ttk.Entry(root, width=10)
        self.empresa_entry.grid(row=3, column=2, padx=125, sticky="w")

        self.conta_label = tk.Label(root, text="Conta", font=("Roboto", 11), bg='#313131', fg="white", anchor='w')
        self.conta_label.grid(row=4, column=2, padx=76, sticky="w")
        self.conta_entry = ttk.Entry(root, width=10)
        self.conta_entry.grid(row=4, column=2, padx=125, sticky="w")

        self.centro_custo_label = tk.Label(root, text="C/Custo", font=("Roboto", 11), bg='#313131', fg="white", anchor='w')
        self.centro_custo_label.grid(row=5, column=2, padx=61, sticky="w")
        self.centro_custo_entry = ttk.Entry(root, width=10)
        self.centro_custo_entry.grid(row=5, column=2, padx=125, sticky="w")

        self.banco_label = tk.Label(root, text="Banco", font=("Roboto", 11), bg='#313131', fg="white", anchor='w')
        self.banco_label.grid(row=4, column=2, padx=1, sticky="e")
        self.banco_entry = ttk.Entry(root, width=20)
        self.banco_entry.grid(row=4, column=3, padx=1, sticky="w")
        
        self.agencia_conta_label = tk.Label(root, text="Agência/Conta", font=("Roboto", 11), bg='#313131', fg="white", anchor='w')
        self.agencia_conta_label.grid(row=5, column=2, padx=1, sticky="e")
        self.agencia_conta_entry = ttk.Entry(root, width=20)
        self.agencia_conta_entry.grid(row=5, column=3, padx=1, sticky="w")

        self.saldo_final_contabil_label = tk.Label(root, text="Saldo Final Contábil", font=("Roboto", 11), bg='#313131', fg="white", anchor='w')
        self.saldo_final_contabil_label.grid(row=4, column=3, columnspan=4, padx=98, sticky="e")
        self.saldo_final_contabil_entry = ttk.Entry(root, width=10)
        self.saldo_final_contabil_entry.grid(row=4, column=3, columnspan=4, padx=10, sticky="e")

        self.diferenca_extrato_bancario_label = tk.Label(root, text="Diferença c/Extrato Bancário", font=("Roboto", 11), bg='#313131', fg="white", anchor='w')
        self.diferenca_extrato_bancario_label.grid(row=5, column=3, columnspan=4, padx=98, sticky="e")
        self.diferenca_extrato_bancario_entry = ttk.Entry(root, width=10)
        self.diferenca_extrato_bancario_entry.grid(row=5, column=3, columnspan=4, padx=10, sticky="e")

        self.linhadeajuda_label = tk.Label(root, text="Coluna 0", fg='#313131', font=("Roboto", 10), bg='#313131')
        self.linhadeajuda_label.grid(row=1, column=0)
        self.linhadeajuda_label = tk.Label(root, text="Coluna 1", fg='#313131', font=("Roboto", 10), bg='#313131')
        self.linhadeajuda_label.grid(row=1, column=1)
        self.linhadeajuda_label = tk.Label(root, text="Coluna 2", fg='#313131', font=("Roboto", 10), bg='#313131')
        self.linhadeajuda_label.grid(row=1, column=2)
        self.linhadeajuda_label = tk.Label(root, text="Coluna 3", fg='#313131', font=("Roboto", 10), bg='#313131')
        self.linhadeajuda_label.grid(row=1, column=3)
        self.linhadeajuda_label = tk.Label(root, text="Coluna 4", fg='#313131', font=("Roboto", 10), bg='#313131')
        self.linhadeajuda_label.grid(row=1, column=4)
        self.linhadeajuda_label = tk.Label(root, text="Coluna 5", fg='#313131', font=("Roboto", 10), bg='#313131')
        self.linhadeajuda_label.grid(row=1, column=5)

        self.busca_btn = ttk.Button(root, text="Buscar", command=self.buscar_termo_treeview, width=6)
        self.busca_btn.grid(row=2, column=2, padx=55, sticky="w")
        self.busca_entry = ttk.Entry(root, font=("Roboto", 10), width=10)
        self.busca_entry.grid(row=2, column=2, padx=125, sticky="w")

        #* -------------------- TREEVIEW PARA EXIBIR OS DADOS IMPORTADOS -------------------- #
        self.frame_tree = tk.Frame(root)
        self.frame_tree.grid(row=7, column=0, columnspan=6, pady=5, padx=10, sticky="nsew")

        self.tree = ttk.Treeview(self.frame_tree, columns=( #? cria a Treeview dentro do frame
            "DataLEB", "DescricaoLEB", "NDocLEB", "ValorLEB", "SaldoLEB",
            "LancamentoLC", "DataLC", "DebitoLC", "D-C/CLC", "CreditoLC",
            "C-C/CLC", "CNPJLC", "HistoricoLC", "ValorLC"
        ), show="headings")

        self.scrollbar_y = ttk.Scrollbar(self.frame_tree, orient="vertical", command=self.tree.yview) #? cria a scrollbar vertical
        self.scrollbar_y.pack(side="right", fill="y")

        self.tree.configure(yscrollcommand=self.scrollbar_y.set) #? configura a treeview para usar a scrollbar

        self.tree.pack(side="left", fill="both", expand=True) #? empacota a treeview
        
        self.tree.tag_configure('negativo', foreground='red') #? torna os valores negativos em vermelho

        self.tree.tag_configure('linha_par', background='#313131') #? cinza
        self.tree.tag_configure('linha_impar', background='#272727') #? branco

        self.tree.tag_configure('destaque', background='#00385B') #? amarelo claro

        #* -------------------- DEFINIR CABEÇALHOS DAS COLUNAS ------------------- #
        self.tree.heading("DataLEB", text="Data")
        self.tree.heading("DescricaoLEB", text="Descrição")
        self.tree.heading("NDocLEB", text="N° Doc")
        self.tree.heading("ValorLEB", text="Valor")
        self.tree.heading("SaldoLEB", text="Saldo")
        self.tree.heading("LancamentoLC", text="Lançamento")
        self.tree.heading("DataLC", text="Data")
        self.tree.heading("DebitoLC", text="Débito")
        self.tree.heading("D-C/CLC", text="D-C/C")
        self.tree.heading("CreditoLC", text="Crédito")
        self.tree.heading("C-C/CLC", text="C-C/C")
        self.tree.heading("CNPJLC", text="CNPJ")
        self.tree.heading("HistoricoLC", text="Histórico")
        self.tree.heading("ValorLC", text="Valor")

        #* -------------------- AJUSTAR O TAMANHO DAS COLUNAS -------------------- #
        self.tree.column("DataLEB", width=60)
        self.tree.column("DescricaoLEB", width=150)
        self.tree.column("NDocLEB", width=60)
        self.tree.column("ValorLEB", width=60)
        self.tree.column("SaldoLEB", width=60)
        self.tree.column("LancamentoLC", width=150)
        self.tree.column("DataLC", width=60)
        self.tree.column("DebitoLC", width=60)
        self.tree.column("D-C/CLC", width=50)
        self.tree.column("CreditoLC", width=60)
        self.tree.column("C-C/CLC", width=50)
        self.tree.column("CNPJLC", width=80)
        self.tree.column("HistoricoLC", width=100)
        self.tree.column("ValorLC", width=80)

        #* -------------------- TORNAR A TREEVIEW EXPANSÍVEL -------------------- #
        root.grid_rowconfigure(7, weight=1)
        root.grid_columnconfigure(0, weight=1)
        root.grid_columnconfigure(1, weight=1)
        root.grid_columnconfigure(2, weight=1)
        root.grid_columnconfigure(3, weight=1)
        root.grid_columnconfigure(4, weight=1)
        root.grid_columnconfigure(5, weight=1)
        
        self.tree.bind("<Double-1>", self.editar_celula_treeview)

        self.total_linhas_label = tk.Label(root, text="Linhas Importadas: 0", font=("Roboto", 11), bg='#313131')
        self.total_linhas_label.grid(row=5, column=0, padx=10, sticky="w")

    

    def buscar_termo_treeview(self):
        termo = self.busca_entry.get().strip().lower()

        for item in self.tree.get_children():
            valores = self.tree.item(item, 'values')

            if any(termo in str(valor).lower() for valor in valores):
                self.tree.item(item, tags=('destaque',))
            else:
                index = self.tree.index(item)
                tag = 'linha_par' if index % 2 == 0 else 'linha_par'
                self.tree.item(item, tags=(tag,))

    def mostrar_selecao_conta(self): #? cria janela toplevel
        janela_selecao_conta = tk.Toplevel(self.root)
        app = TelaSelecaoConta(janela_selecao_conta, self.abrir_explorador_arquivos)
        self.janelas_filhas.append(janela_selecao_conta) #? adiciona janela filha à lista

    def abrir_tela_nova_empresa(self):
        if not hasattr(self, 'tela_nova_empresa') or not self.tela_nova_empresa.winfo_exists():
            self.tela_nova_empresa = tk.Toplevel(self.root)
            app = TelaNovaEmpresa(self.tela_nova_empresa, self.salvar_nova_empresa)
            self.janelas_filhas.append(self.tela_nova_empresa) #? adiciona janela filha à lista

    def abrir_tela_nova_conta(self):
        if not hasattr(self, 'tela_nova_conta') or not self.tela_nova_conta.winfo_exists():
            self.tela_nova_conta = tk.Toplevel(self.root)
            app = TelaNovaConta(self.tela_nova_conta, self.salvar_nova_conta)
            self.janelas_filhas.append(self.tela_nova_conta) #? adiciona janela filha à lista

    def fechar_janela(self):
        for janela in self.janelas_filhas: #? fecha todas as janelas filhas
            if janela.winfo_exists():
                janela.destroy()
        self.root.destroy() #? fecha a janela principal

    def abrir_explorador_arquivos(self, empresa, conta, arquivo, respostas_safra=None):
        try:
            conta_ativo, banco, agencia, conta_bancaria, empresa_codigo = conta.split(",")
        except:
            messagebox.showerror("Erro", "O formato da conta está incorreto.")
            return
        
        self.empresa_entry.delete(0, tk.END)
        self.empresa_entry.insert(0, empresa_codigo.strip())
        self.conta_entry.delete(0, tk.END)
        self.conta_entry.insert(0, conta_ativo.strip())
        self.banco_entry.delete(0, tk.END)
        self.banco_entry.insert(0, banco.strip())

        self.agencia_conta_entry.delete(0, tk.END)
        self.agencia_conta_entry.insert(0, f"{agencia.strip()}/{conta_bancaria.strip()}")

        self.detectar_banco(arquivo, respostas_safra)

    def detectar_banco(self, nome_arquivo, respostas_safra=None):
        bancos = [
            "Bco.Brasil", "Banco do Brasil", "BANCO DO BRASIL", "BB", "BRASIL", "Brasil",
            "Bco.Inter", "Banco Inter", "BANCO INTER", "Inter", "INTER",
            "Bco.Caixa", "Banco Caixa Eletrônica", "BANCO CAIXA ELETRÔNICA", "Caixa Eletrônica", "Caixa", "CAIXA", "CX",
            "Bco.Bradesco", "Banco Bradesco", "BANCO BRADESCO", "Bradesco", "EXTRATO BRADESCO",
            "Bco.Grafeno", "Banco Grafeno", "BANCO GRAFENO", "GRAFENO", "Grafeno",
            "Bco.Pagseguro", "Banco Pagseguro", "BANCO PAGSEGURO", "PAGSEGURO", "Pagseguro",
            "Bco.C6Bank", "Banco C6 Bank", "C6 Bank", "BANCO C6 BANK", "C6 BANK", "C6BANK", "C6",
            "Bco.Santander", "Banco Santander", "BANCO SANTANDER", "SANTANDER", "Santander",
            "Bco.HSBC", "Banco HSBC", "BANCO HSBC", "HSBC",
            "Bco.Safra", "Banco Safra", "BANCO SAFRA", "SAFRA", "Safra",
            "Bco.Suisse", "Banco Suisse", "Banco Credit Suisse", "Credit Suisse", "CREDIT SUISSE", "SUISSE", "BANCO SUISSE",
            "Bco.Daycoval", "Banco Daycoval", "BANCO DAYCOVAL", "DAYCOVAL", "Daycoval",
            "Bco.Itaú", "Bco.Itau", "Banco Itaú", "BANCO ITAÚ", "ITAÚ", "Itaú", "Itau", "ITAU",
        ]
    
        for banco in bancos:
            if banco in nome_arquivo:
                logging.debug(f"Banco detectado: {banco}")
                self.executar_acao_para_banco(banco, nome_arquivo, respostas_safra)
                break
        else:
            logging.error("Nenhum banco detectado no nome do arquivo.")

    def copiar_coluna(self, coluna_origem, coluna_destino):
        for item in self.tree.get_children():
            values = self.tree.item(item, 'values')

            valor_origem = values[self.tree["columns"].index(coluna_origem)]
            self.tree.set(item, coluna_destino, valor_origem)

        #* -------------------- ETAPA DA CLISSIFICAÇÃO -------------------- #
    def corrigir_valor(self, valor_str):
        try:
            if isinstance(valor_str, str):
                valor_str = valor_str.replace('.', '').replace(',', '.')
                return float(valor_str)
        except Exception as e:
            logging.warning(f"[ERRO] Falha ao corrigir valor '{valor_str}': {e}")
            return 0.0

    def classificar_dados(self):
        logging.info("\n=== PROCESSO DE CLASSIFICAÇÃO ===")
        conta_bancaria = self.conta_entry.get()
        incluir_banco = messagebox.askyesno("Classificar Dados", "Deseja incluir a identificação do banco no histórico contábil?")
    
        conta_ativo = None
        try:
            if os.path.exists('contas.csv'):
                df_contas = pd.read_csv('contas.csv')
                filtro = (df_contas['empresa'] == self.empresa_entry.get()) & (df_contas['banco'] == self.banco_entry.get())
                conta_encontrada = df_contas[filtro]
                if not conta_encontrada.empty:
                    conta_ativo = str(conta_encontrada.iloc[0]['Conta Ativo']).strip()
                    logging.debug(f"Conta ativo encontrada: {conta_ativo}")
        except Exception as e:
            logging.warning(f"Erro ao carregar contas.csv: {e}")

        desc_idx = self.tree["columns"].index("DescricaoLEB")
        hist_idx = self.tree["columns"].index("HistoricoLC")

        colunas = list(self.tree["columns"])
        logging.debug("Ordem das colunas:", colunas)
        logging.debug("Descrição index:", desc_idx)
        logging.debug("Histórico index:", hist_idx)

        if incluir_banco:
            banco_nome = self.banco_entry.get().strip()
            agencia_conta = self.agencia_conta_entry.get().strip()
            if '/' in agencia_conta:
                agencia_nome, conta_nome = agencia_conta.split('/', 1)
            else:
                agencia_nome = agencia_conta
                conta_nome = self.conta_entry.get().strip()

        for item in self.tree.get_children():
            logging.debug(f"ITEM: {item}")
            logging.debug(f"Valores: {self.tree.item(item, 'values')}")
            values = list(self.tree.item(item, 'values'))

        # Copia DescricaoLEB → HistoricoLC
            values[hist_idx] = values[desc_idx]

        # Adiciona dados bancários ao final, se solicitado
            if incluir_banco and values[hist_idx]:
                values[hist_idx] = f"{values[hist_idx]} - Bco.{banco_nome} Ag.{agencia_nome} CC.{conta_nome}"

            self.tree.item(item, values=values)

        lancamento_idx =  self.tree["columns"].index("LancamentoLC")
        for i, item in enumerate(self.tree.get_children(), start=1):
            values = list(self.tree.item(item, 'values'))
            values[lancamento_idx] = str(i)
            self.tree.item(item, values=values)

    # Classificação usando fuzzy matching
        df_descricoes_normalizadas = self.df_banco_dados['Descricao'].apply(
            lambda x: unidecode(str(x).strip().upper())
        )
        df_banco_tipo_normalizado = self.df_banco_dados['Tipo'].apply(lambda x: str(x).strip().upper())

        for idx, item in enumerate(self.tree.get_children(), start=1):
            values = self.tree.item(item, 'values')
            hist_idx = self.tree["columns"].index("HistoricoLC")
            valor_idx = self.tree["columns"].index("ValorLEB")

            descricao = values[hist_idx]
            valor_raw = values[valor_idx]

            try:
                valor = self.corrigir_valor(valor_raw)
                tipo = "D" if valor < 0 else "C"
            except Exception as e:
                logging.warning(f"[ERRO] Falha ao converter valor: '{valor_raw}' -> {e}")
                continue

            descricao_normalizada = unidecode(str(descricao).strip().upper())

            df_filtrado = self.df_banco_dados[self.df_banco_dados["Tipo"].str.upper() == tipo]
            df_descricoes_normalizadas = df_filtrado["Descricao"].apply(lambda x: unidecode(str(x).strip().upper()))

            melhores_matches = []
            for i, descricao_banco in enumerate(df_descricoes_normalizadas):
                score = fuzz.partial_ratio(descricao_normalizada, descricao_banco)
                if score >= 80:
                    melhores_matches.append((descricao_banco, score, i))

            if melhores_matches:
                melhores_matches.sort(key=lambda x: (x[1], len(x[0])), reverse=True)
                descricao_mais_parecida, score_usado, indice_match = melhores_matches[0]
                correspondencia = df_filtrado.iloc[indice_match]
            
                debito = correspondencia.iloc[6]
                credito = correspondencia.iloc[9]

                if "#BCO" in str(debito):
                    debito = conta_bancaria
                if "#BCO" in str(credito):
                    credito = conta_bancaria

                logging.debug(f"📌 Linha {idx} 📄 Treeview: {descricao}")
                logging.debug(f"🎒 CSV: {descricao_mais_parecida}")
                logging.debug(f"🔍 Tipo: {tipo} 🔴 Débito: {debito} 🟢 Crédito: {credito} 🎯 Similaridade: {score}")
                logging.debug("-" * 60)

                novos_valores = list(values)
                novos_valores[7] = debito   # DebitoLC
                novos_valores[9] = credito  # CreditoLC
                self.tree.item(item, values=novos_valores)

        self.copiar_coluna("DataLEB", "DataLC")
        self.copiar_coluna("ValorLEB", "ValorLC")
        messagebox.showinfo("Classificar Dados", "Dados classificados com sucesso!")

    def processar_pdf_bancario_bruto(self, linhas):
        transacoes = []
        ultima_data = None
        transacao_atual = {"data": None, "descricao": "", "dcto": "", "credito": "", "debito": "", "saldo": ""}

        data_regex = re.compile(r'^\d{2}/\d{2}/\d{4}')
        valores_regex = re.compile(r'(\d+)\s+([-\d.,]+)?\s+([-\d.,]+)?$')

        for i, linha in enumerate(linhas):
            linha = linha.strip()
            if not linha:
                continue

            logging.debug(f"🔹 [{i}] Lendo linha: {linha}")

            if "SALDO ANTERIOR" in linha.upper():
                logging.debug("⏭ Ignorando linha de saldo anterior.")
                continue

            if data_regex.match(linha):
                if transacao_atual["descricao"]:
                    logging.debug(f"🔄 Fechando transação: {transacao_atual}")
                    transacoes.append(transacao_atual)
                    transacao_atual = {"data": None, "descricao": "", "dcto": "", "credito": "", "debito": "", "saldo": ""}

                partes = linha.split(" ", 1)
                ultima_data = partes[0]
                transacao_atual["data"] = ultima_data
                transacao_atual["descricao"] = partes[1] if len(partes) > 1 else ""

            elif valores_regex.search(linha):
                match = valores_regex.search(linha)
                resto = linha[:match.start()].strip()

                if not transacao_atual["data"]:
                    transacao_atual["data"] = ultima_data

                transacao_atual["descricao"] += " " + resto
                grupos = match.groups()
                transacao_atual["dcto"] = grupos[0]

                valor1 = grupos[1]
                valor2 = grupos[2]

            # Trata crédito e débito
                credito = valor1 if valor1 and not '-' in valor1 else ''
                debito = valor1 if valor1 and '-' in valor1 else ''

                if not debito and valor2 and '-' in valor2:
                    debito = valor2
                elif not credito and valor2 and '-' not in valor2:
                    credito = valor2

                transacao_atual["credito"] = credito or ''
                transacao_atual["debito"] = debito or ''
                transacao_atual["saldo"] = valor2 or ''

                logging.info(f"✅ Transação processada: {transacao_atual}")
                transacoes.append(transacao_atual)
                transacao_atual = {"data": None, "descricao": "", "dcto": "", "credito": "", "debito": "", "saldo": ""}

            else:
                transacao_atual["descricao"] += " " + linha

        if transacao_atual["descricao"]:
            transacao_atual["data"] = transacao_atual["data"] or ultima_data
            logging.debug(f"🔚 Fechando última transação: {transacao_atual}")
            transacoes.append(transacao_atual)

        return transacoes

    def executar_acao_para_banco(self, nome_banco, arquivo, respostas_safra=None):
        logging.info(f"Banco detectado: {nome_banco}")
        acoes = {
            "Bco.Bradesco": self.acao_bradesco, "Banco Bradesco": self.acao_bradesco, "BANCO BRADESCO": self.acao_bradesco, "Bradesco": self.acao_bradesco, "EXTRATO BRADESCO": self.acao_bradesco,
            "Bco.Safra": self.acao_safra, "Banco Safra": self.acao_safra, "BANCO SAFRA": self.acao_safra, "SAFRA": self.acao_safra, "Safra": self.acao_safra, "EXTRATO SAFRA": self.acao_safra,
            "Bco.Daycoval": self.acao_daycoval, "Banco Daycoval": self.acao_daycoval, "BANCO DAYCOVAL": self.acao_daycoval, "DAYCOVAL": self.acao_daycoval, "Daycoval": self.acao_daycoval, "EXTRATO DAYCOVAL": self.acao_daycoval,
            "Bco.Itaú": self.acao_itau, "Bco.Itau": self.acao_itau, "Banco Itaú": self.acao_itau, "BANCO ITAÚ": self.acao_itau, "ITAÚ": self.acao_itau, "Itaú": self.acao_itau, "EXTRATO ITAU": self.acao_itau, "EXTRATO ITAÚ": self.acao_itau,
            "Bco.Brasil": self.acao_brasil, "Banco do Brasil": self.acao_brasil, "BANCO DO BRASIL": self.acao_brasil, "BB": self.acao_brasil, "BRASIL": self.acao_brasil, "Brasil": self.acao_brasil, "EXTRATO BB": self.acao_brasil,
            "Bco.Santander": self.acao_santander, "Banco Santander": self.acao_santander, "BANCO SANTANDER": self.acao_santander, "SANTANDER": self.acao_santander, "Santander": self.acao_santander, "EXTRATO SANTANDER": self.acao_santander,
            "Bco.Inter": self.acao_inter, "Banco Inter": "acao_inter", "BANCO INTER": self.acao_inter, "Inter": self.acao_inter, "INTER": self.acao_inter, "EXTRATO INTER": self.acao_inter,
            "Bco.Caixa": self.acao_caixa, "Banco Caixa Eletrônica": self.acao_caixa, "BANCO CAIXA ELETRÔNICA": self.acao_caixa, "Caixa Eletrônica": self.acao_caixa, "Caixa": self.acao_caixa, "EXTRATO CAIXA": self.acao_caixa, "CAIXA": self.acao_caixa, "CX": self.acao_caixa,
            "Bco.Grafeno": self.acao_grafeno, "Banco Grafeno": self.acao_grafeno, "BANCO GRAFENO": self.acao_grafeno, "GRAFENO": self.acao_grafeno, "Grafeno": self.acao_grafeno, "EXTRATO GRAFENO": self.acao_grafeno,
            "Bco.Pagseguro": self.acao_pagseguro, "Banco Pagseguro": self.acao_pagseguro, "BANCO PAGSEGURO": self.acao_pagseguro, "PAGSEGURO": self.acao_pagseguro, "Pagseguro": self.acao_pagseguro, "EXTRATO PAGSEGURO": self.acao_pagseguro,
            "Bco.C6Bank": self.acao_c6bank, "Banco C6 Bank": self.acao_c6bank, "C6 Bank": self.acao_c6bank, "BANCO C6 BANK": self.acao_c6bank, "C6 BANK": self.acao_c6bank, "C6BANK": self.acao_c6bank, "C6": self.acao_c6bank, "c6": self.acao_c6bank, "EXTRATO C6": self.acao_c6bank,
            "Bco.Santander": self.acao_santander, "Banco Santander": self.acao_santander, "BANCO SANTANDER": self.acao_santander, "SANTANDER": self.acao_santander, "Santander": self.acao_santander, "EXTRATO SANTANDER": self.acao_santander,
            "Bco.HSBC": self.acao_hsbc, "Banco HSBC": self.acao_hsbc, "BANCO HSBC": self.acao_hsbc, "HSBC": self.acao_hsbc, "EXTRATO HSBC": self.acao_hsbc,
            "Bco.Suisse": self.acao_suisse, "Banco Suisse": self.acao_suisse, "Banco Credit Suisse": self.acao_suisse, "Credit Suisse": self.acao_suisse, "CREDIT SUISSE": self.acao_suisse, "SUISSE": self.acao_suisse, "BANCO SUISSE": self.acao_suisse, "EXTRATO SUISSE": self.acao_suisse,
            "Bco.Daycoval": self.acao_daycoval, "Banco Daycoval": self.acao_daycoval, "BANCO DAYCOVAL": self.acao_daycoval, "DAYCOVAL": self.acao_daycoval, "Daycoval": self.acao_daycoval, "EXTRATO DAYCOVAL": self.acao_daycoval,
            "Bco.Sicredi": self.acao_sicredi, "Banco Sicredi": self.acao_sicredi, "BANCO SICREDI": self.acao_sicredi, "SICREDI": self.acao_sicredi, "Sicredi": self.acao_sicredi, "EXTRATO SICREDI": self.acao_sicredi,
        }

        logging.debug(f"Chaves disponíveis: {list(acoes.keys())}")

        if nome_banco in acoes:
            logging.info(f"Executando ação específica para {nome_banco}.")
            if nome_banco in ["Bco.Safra", "Banco Safra", "BANCO SAFRA", "SAFRA", "Safra", "EXTRATO SAFRA"]:
                acoes[nome_banco](arquivo, respostas_safra)
            else:
                acoes[nome_banco](arquivo)
        else:
            logging.info(f"Executando ação padrão para {nome_banco}.")

    def acao_bradesco(self, arquivo):
            logging.info("\n=== INÍCIO DO PROCESSAMENTO: BANCO BRADESCO ===")
            logging.debug(f"Arquivo recebido: {arquivo}")
            def formatar_valor_brasileiro(valor):
                try:
                    return locale.format_string("%.2f", float(valor), grouping=True)
                except:
                    return valor
            try:
                extensao = arquivo.lower().split('.')[-1]
                logging.debug(f"Extensão detectada: {extensao}")
                dados_importados = []
                saldo_final_calculado = 0  #? inicializa a variável aqui

                if extensao == 'pdf': #! SE O ARQUIVO É PDF
                    logging.debug("\n=== PROCESSANDO ARQUIVO PDF ===")
                    try:
                        import pdfplumber
                        with pdfplumber.open(arquivo) as pdf:
                            linhas = []
                            capturar = False
                            for pagina in pdf.pages:
                                texto = pagina.extract_text()
                                if texto:
                                    for linha in texto.split("\n"):
                                        if not capturar and "SALDO ANTERIOR" in linha.upper():
                                            capturar = True
                                            logging.debug(f"Encontrado 'SALDO ANTERIOR', iniciando leitura a partir daqui...")
                                            continue
                                        if capturar:
                                            linhas.append(linha.strip())

                        logging.info(f"📄 Total de linhas extraídas: {len(linhas)}")
                        transacoes = self.processar_pdf_bancario_bruto(linhas)

                        for i, t in enumerate(transacoes):
                            logging.debug(f"📤 Transação {i+1}: {t}")
                            self.tree.insert("", "end", values=[
                                t["data"], t["descricao"].strip(), t["dcto"],
                                t["credito"] or ("-" + t["debito"] if t["debito"] else ""),  # valor
                                t["saldo"], "", "", "", "", "", "", "", "", ""
                            ])

                        self.atualizar_total_linhas_importadas()
                        messagebox.showinfo("Sucesso", f"{len(transacoes)} transações importadas do PDF!")

                    except Exception as e:
                        logging.critical(f"❌ Erro ao processar PDF: {e}")
                        traceback.print_exc()
                        messagebox.showerror("Erro", f"Erro ao processar o PDF:\n\n{str(e)}")
                    return
                
                elif extensao == 'xls': #! SE O ARQUIVO É XLS
                    logging.info("\n=== PROCESSANDO ARQUIVO XLS ===")
                    wb = xlrd.open_workbook(arquivo)
                    sheet = wb.sheet_by_index(0)
                    logging.debug(f"Planilha aberta: {sheet.name}")
                    logging.debug(f"Dimensões: {sheet.nrows} linhas x {sheet.ncols} colunas")
                    
                    logging.debug("\nBuscando saldo inicial...") #? lê o saldo inicial (F10)
                    saldo_inicial = sheet.cell_value(9, 5)
                    logging.debug(f"Valor bruto encontrado em F10: {saldo_inicial}")
                    logging.debug(f"Tipo do valor: {type(saldo_inicial)}")
                    
                    if isinstance(saldo_inicial, str): #? formata e confirma saldo inicial
                        logging.debug("Convertendo saldo inicial de string para float...")
                        saldo_inicial = float(saldo_inicial.replace(".", "").replace(",", "."))
                    saldo_inicial_frmt = locale.format_string("%.2f", saldo_inicial, grouping=True)
                    logging.debug(f"Saldo inicial formatado: R${saldo_inicial_frmt}")
                    
                    resposta = messagebox.askyesno("Confirmação de saldo", f"O saldo inicial é de R${saldo_inicial_frmt}?")

                    if not resposta:
                        logging.info("Usuário não confirmou o saldo inicial, solicitando entrada manual.")
                        saldo_inicial_manual = simpledialog.askstring("Entrada de Saldo", "Insira o saldo inicial correto:")
                        if saldo_inicial_manual:
                            try:
                                saldo_inicial = float(saldo_inicial_manual.replace(".", "").replace(",", "."))
                                saldo_inicial_frmt = locale.format_string("%.2f", saldo_inicial, grouping=True)
                            except ValueError:
                                messagebox.showerror("Erro", "Valor de saldo inicial inválido.")
                                return
                        else:
                            messagebox.showinfo("Aviso", "Processo cancelado pelo usuário.")
                            return
                        
                    logging.debug("Atualizando campo de saldo inicial na interface...")
                    self.saldo_inicial_entry.delete(0, tk.END)
                    self.saldo_inicial_entry.insert(0, saldo_inicial_frmt)
                    
                    saldo_final_calculado = saldo_inicial #? processa as linhas
                    
                    logging.info("\n=== INICIANDO PROCESSAMENTO DAS LINHAS ===")
                    logging.debug(f"Total de linhas na planilha: {sheet.nrows}")
                    
                    for row in range(10, sheet.nrows):
                        try:
                            logging.debug(f"\nProcessando linha {row+1}:")
                            data = sheet.cell_value(row, 0)
                            logging.debug(f"Data encontrada: {data} (tipo: {type(data)})")
                            
                            if not data:
                                logging.debug("Linha vazia, pulando...")
                                continue
                            if isinstance(data, str) and "total" in data.lower():
                                logging.debug("Encontrada linha de total, parando processamento...")
                                break
                                
                            historico = sheet.cell_value(row, 1)
                            num_doc = sheet.cell_value(row, 2)
                            credito = sheet.cell_value(row, 3)
                            debito = sheet.cell_value(row, 4)
                            saldo = sheet.cell_value(row, 5)
                            
                            logging.info(f"Valores lidos:")
                            logging.info(f"  Histórico: {historico}, Nº Doc: {num_doc}, Crédito: {credito}, Débito: {debito}, Saldo: {saldo}")
                            
                            def converter_para_float(valor):
                                if valor is None or valor == "":
                                    return 0.0
                                if isinstance(valor, str):
                                    valor = valor.replace(".", "").replace(",", ".")
                                try:
                                    return float(valor)
                                except ValueError:
                                    return 0.0
                            
                            valor_credito = converter_para_float(credito)
                            valor_debito = converter_para_float(debito)
                            valor_total = valor_credito + valor_debito
                            logging.info(f"Valor total calculado: {valor_total}")
                            
                            if isinstance(data, float): #? formata a data se for um número
                                logging.debug("Convertendo data de float para string...")
                                data = xlrd.xldate_as_datetime(data, wb.datemode).strftime('%d/%m/%Y')
                                logging.debug(f"Data convertida: {data}")
                            
                            logging.debug("Adicionando linha aos dados importados...")
                            dados_importados.append([
                                data, historico, num_doc, valor_total, saldo,
                                "", "", "", "", "", "", "", "", ""
                            ])
                            
                            saldo_final_calculado += valor_total
                            logging.debug(f"Novo saldo calculado: {saldo_final_calculado}")
                            
                        except Exception as e:
                            logging.warning(f"ERRO ao processar linha {row+1}:")
                            logging.error(f"Detalhes do erro: {str(e)}")
                            traceback.print_exc()
                            continue
                    
                elif extensao == "xlsx":  #! SE O ARQUIVO É XLSX
                    logging.info("\n=== PROCESSANDO ARQUIVO XLSX ===")
                    wb = openpyxl.load_workbook(arquivo, data_only=True)
                    sheet = wb.active
                    logging.debug(f"Planilha ativa: {sheet.title}")
                    
                    logging.debug("\nBuscando saldo inicial...") #? lê o saldo inicial (F10)
                    saldo_inicial_celula = sheet['F10'].value
                    logging.debug(f"Valor bruto encontrado em F10: {saldo_inicial_celula}")
                    
                    if isinstance(saldo_inicial_celula, str):
                        saldo_inicial = float(saldo_inicial_celula.replace(".", "").replace(",", "."))
                    else:
                        saldo_inicial = float(saldo_inicial_celula)
                        
                    saldo_inicial_frmt = locale.format_string("%.2f", saldo_inicial, grouping=True)
                    logging.debug(f"Saldo inicial formatado: R${saldo_inicial_frmt}")
                    
                    resposta = messagebox.askyesno("Confirmação de saldo", 
                                                 f"O saldo inicial é de R${saldo_inicial_frmt}?")
                    if not resposta:
                        logging.debug("Usuário não confirmou o saldo inicial. Abortando...")
                        return
                        
                    self.saldo_inicial_entry.delete(0, tk.END)
                    self.saldo_inicial_entry.insert(0, saldo_inicial_frmt)
                    
                    #? inicializa o saldo final calculado com o saldo inicial
                    saldo_final_calculado = saldo_inicial
                    
                    logging.info("\n=== INICIANDO PROCESSAMENTO DAS LINHAS ===")
                    for row in range(11, sheet.max_row + 1):
                        try:
                            logging.debug(f"\nProcessando linha {row}:")
                            data = sheet.cell(row=row, column=1).value
                            logging.debug(f"Data encontrada: {data}")
                            
                            if not data:
                                logging.debug("Linha vazia, pulando...")
                                continue
                            if isinstance(data, str) and "total" in data.lower():
                                logging.debug("Encontrada linha de total, parando processamento...")
                                break
                                
                            historico = sheet.cell(row=row, column=2).value
                            num_doc = sheet.cell(row=row, column=3).value
                            credito = sheet.cell(row=row, column=4).value
                            debito = sheet.cell(row=row, column=5).value
                            saldo = sheet.cell(row=row, column=6).value
                            
                            logging.info(f"Valores lidos:")
                            logging.info(f"  Histórico: {historico}, Nº Doc: {num_doc}, Crédito: {credito}, Débito: {debito}, Saldo: {saldo}")
                            
                            def converter_para_float(valor):
                                if valor is None or valor == "":
                                    return 0.0
                                if isinstance(valor, str):
                                    valor = valor.replace(".", "").replace(",", ".")
                                try:
                                    return float(valor)
                                except ValueError:
                                    return 0.0
                            
                            valor_credito = converter_para_float(credito)
                            valor_debito = converter_para_float(debito)
                            valor_total = valor_credito + valor_debito
                            logging.info(f"Valor total calculado: {valor_total}")
                            
                            dados_importados.append([
                                data, historico, num_doc, valor_total, saldo,
                                "", "", "", "", "", "", "", "", ""
                            ])
                            
                            saldo_final_calculado += valor_total
                            logging.info(f"Novo saldo calculado: {saldo_final_calculado}")
                            
                        except Exception as e:
                            logging.warning(f"ERRO ao processar linha {row}:")
                            logging.error(f"Detalhes do erro: {str(e)}")
                            continue

                else:
                    logging.critical("Extensão não suportada.")
                    messagebox.showerror("Erro", "Formato de arquivo não suportado.")
                
                logging.info("\n=== ATUALIZANDO INTERFACE ===")
                logging.debug("Formatando saldo final...")
                saldo_final_calculado_frmt = locale.format_string("%.2f", saldo_final_calculado, grouping=True)
                logging.debug(f"Saldo final formatado: R${saldo_final_calculado_frmt}")
                
                logging.debug("Atualizando campo de saldo final...")
                self.saldo_final_calculado_entry.delete(0, tk.END)
                self.saldo_final_calculado_entry.insert(0, saldo_final_calculado_frmt)
                
                logging.debug("\nLimpando Treeview...")
                for i in self.tree.get_children():
                    self.tree.delete(i)
                    
                logging.debug("Inserindo dados na Treeview...")
                logging.debug(f"Total de registros a inserir: {len(dados_importados)}")

                for i, dados in enumerate(dados_importados):
                    tag = 'linha_par' if i % 2 == 0 else 'linha_impar'

                    dados[3] = formatar_valor_brasileiro(dados[3])

                    self.tree.insert("", "end", values=dados, tags=(tag,))

                self.atualizar_total_linhas_importadas()
                    
                logging.info("\n=== PROCESSAMENTO CONCLUÍDO COM SUCESSO ===")
                logging.info(f"Total de linhas processadas: {len(dados_importados)}")
                
            except Exception as e:
                logging.critical("\n=== ERRO FATAL ===")
                logging.critical(f"Erro: {str(e)}")
                logging.critical("Stack trace:")
                traceback.print_exc()
                messagebox.showerror("Erro", 
                    "Erro ao processar o arquivo. Verifique se:\n\n" +
                    "1. O arquivo está no formato correto\n" +
                    "2. O arquivo não está em modo de exibição protegida\n" +
                    "3. O arquivo está fechado no Excel\n\n" +
                    f"Erro: {str(e)}")
                return

    def acao_safra(self, arquivo, respostas_safra):
        logging.info("\n=== INÍCIO DO PROCESSAMENTO: SAFRA ===")
        def formatar_valor_brasileiro(valor):
            try:
                return locale.format_string("%.2f", float(valor), grouping=True)
            except:
                return valor

        if respostas_safra:
            novo_formato = respostas_safra.get("novo_formato", False)
            conta_vinculada = respostas_safra.get("conta_vinculada", False)
            extensao = arquivo.lower().split('.')[-1]

            if extensao in ["xls", "xlsx"]:
                if novo_formato:
                    if conta_vinculada: #! LÓGICA PARA ARQUIVO XLS/XLSX, NOVO FORMATO E CONTA VINCULADA
                        logging.info("Processando XLS/XLSX, novo formato no Safra, conta vinculada.")
                        logging.debug(f"Arquivo recebido: {arquivo}")
                        import pdfplumber
                        import re
                        from tkinter import messagebox

                        try:
                            dados_importados = []
                            saldo_final_calculado = 0

                            if extensao == "xls": #! SE O ARQUIVO É XLS
                                import xlrd
                                logging.debug("\n=== PROCESSANDO ARQUIVO XLS ===")
                                logging.debug("Abrindo workbook...")
                                wb = xlrd.open_workbook(arquivo)
                                sheet = wb.sheet_by_index(0)

                                texto_periodo = sheet.cell_value(5, 1)
                                logging.debug(f"Texto do período: {texto_periodo}")

                                def extrair_ano(texto):
                                    match = re.search(r'\d{2}/\d{2}/(\d{4})', texto)
                                    if match:
                                        return int(match.group(1))
                                    return datetime.now().year
                                
                                ano_extrato = extrair_ano(texto_periodo)
                                logging.debug(f"Ano extraído do período: {ano_extrato}")

                                meses_pt = {
                                    'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4, 'mai': 5, 'jun': 6,
                                    'jul': 7, 'ago': 8, 'set': 9, 'out': 10, 'nov': 11, 'dez': 12
                                }

                                def converter_data_pt(data_str, ano):
                                    try:
                                        dia, mes_abrev = data_str.strip().split('/')
                                        mes = meses_pt.get(mes_abrev.lower())
                                        if mes:
                                            data = datetime(ano, mes, int(dia))
                                            return data.strftime('%d/%m/%Y')
                                    except Exception as e:
                                        logging.warning(f"Erro ao converter data '{data_str}': {e}")
                                    return data_str

                                logging.debug(f"Planilha aberta: {sheet.name}")
                                logging.debug(f"Dimensões? {sheet.nrows} linhas x {sheet.ncols} colunas")

                                logging.debug("\nBuscando saldo inicial...")
                                saldo_inicial = sheet.cell_value(3, 5)
                                logging.debug(f"Valor bruto encontrado em F4: {saldo_inicial}")
                                logging.debug(f"Tipo do valor: {type(saldo_inicial)}")

                                try:
                                    if isinstance(saldo_inicial, str):
                                        logging.debug("Convertendo saldo inicial de string para float...")
                                        saldo_inicial = float(saldo_inicial.replace(".", "").replace(",", "."))
                                    saldo_inicial_frmt = locale.format_string("%.2f", saldo_inicial, grouping=True)
                                except ValueError:
                                    saldo_inicial_frmt = ""
                                    saldo_inicial = 0.0
                                logging.debug(f"Saldo inicial formatado: R${saldo_inicial_frmt}")

                                resposta = messagebox.askyesno("Confirmação de saldo", f"O saldo inicial é de R${saldo_inicial_frmt}?")

                                if not resposta:
                                    logging.debug("Usuário não confirmou o saldo inicial, solicitando entrada manual.")
                                    saldo_inicial_manual = simpledialog.askstring("Entrada de Saldo", "Insira o saldo inicial correto:")
                                    if saldo_inicial_manual:
                                        try:
                                            saldo_inicial = float(saldo_inicial_manual.replace(".", "").replace(",", "."))
                                            saldo_inicial_frmt = locale.format_string("%.2f", saldo_inicial, grouping=True)
                                        except ValueError:
                                            messagebox.showerror("Erro", "Valor de saldo inicial inválido.")
                                            return
                                    else:
                                        messagebox.showinfo("Aviso", "Processo cancelado pelo usuário.")
                                        return
                    
                                logging.debug("Atualizando campo de saldo inicial na interface...")
                                self.saldo_inicial_entry.delete(0, tk.END)
                                self.saldo_inicial_entry.insert(0, saldo_inicial_frmt)

                                saldo_final_calculado = saldo_inicial

                                logging.info("\n=== INICIANDO PROCESSAMENTO DAS LINHAS ===")
                                logging.debug(f"Total de linhas na planilha: {sheet.nrows}")

                                for row in range(8, sheet.nrows):
                                    try:
                                        logging.debug(f"\nProcessando linha {row+1}:")
                                        data = sheet.cell_value(row, 0)
                                        logging.debug(f"Data encontrada: {data} (tipo: {type(data)})")

                                        historico = sheet.cell_value(row, 1)
                                        num_doc = sheet.cell_value(row, 4)
                                        valor = sheet.cell_value(row, 5)

                                        logging.info(f"Valores lidos:")
                                        logging.info(f"  Histórico: {historico}, N° Doc: {num_doc}, Valor: {valor}")

                                        logging.debug(f"Processando descrições...")
                                        valor_celula = sheet.cell_value(row, 1)
                                        if isinstance(valor_celula, str) and valor_celula.strip() in ["SALDO POUPANCA PLUS", "CONTA VINCULADA"]:
                                            logging.debug("Saldo/conta encontrada, pulando linha...")
                                            continue

                                        def converter_para_float(valor):
                                            logging.debug(f"Tratando valor: {valor} (tipo: {type(valor)})")
                                            if valor is None or valor == "":
                                                logging.debug("Valor vazio, retornando 0.0")
                                                return 0.0
                                            if isinstance(valor, str):
                                                logging.debug("Convertendo string para float...")
                                                valor = valor.replace(".", "").replace(",", ".")
                                            try:
                                                resultado = float(valor)
                                                logging.debug(f"Valor convertido: {resultado}")
                                                return resultado
                                            except ValueError as e:
                                                logging.warning(f"Erro ao converter valor: {e}")
                                                return 0.0

                                        valor_total = converter_para_float(valor)
                                        valor_formatado = formatar_valor_brasileiro(valor_total)
                                        logging.debug(f"Valor total calculado: {valor_formatado}")

                                        if isinstance(data, float):
                                            data = xlrd.xldate_as_datetime(data, wb.datemode).strftime('%d/%m/%Y')
                                        elif isinstance(data, str):
                                            if re.match(r'\d{2}/\d{2}', data):  #? ex: 02/01
                                                try:
                                                    data = datetime.strptime(f"{data}/{ano_extrato}", "%d/%m/%Y").strftime("%d/%m/%Y")
                                                except Exception as e:
                                                    logging.warning(f"Erro ao converter data '{data}': {e}")
                                            elif re.match(r'\d{2}/[a-zA-Z]{3}', data):  #? ex: 02/jan
                                                data = converter_data_pt(data, ano_extrato)
                            
                                        logging.info("Adicionando linha aos dados importados...")
                                        dados_importados.append([
                                            data, historico, num_doc, valor_formatado, "",
                                            "", "", "", "", "", "", "", ""
                                        ])

                                        saldo_final_calculado += valor_total
                                        saldo_final_calculado = round(saldo_final_calculado, 2)
                                        logging.info(f"Novo saldo calculado: {saldo_final_calculado}")
                            
                                    except Exception as e:
                                        logging.info(f"ERRO ao processar linha {row+1}:")
                                        logging.info(f"Detalhes do erro: {str(e)}")
                                        traceback.print_exc()
                                        continue

                            else: #! SE O ARQUIVO É XLSX
                                logging.info("\n=== PROCESSANDO ARQUIVO XLSX ===")
                                logging.info(f"Arquivo recebido: {arquivo}")
                                wb = openpyxl.load_workbook(arquivo, data_only=True)
                                sheet = wb.active
                                logging.info(f"Planilha ativa: {sheet.title}")

                                texto_periodo = sheet["B6"].value
                                logging.info(f"Texto do período: {texto_periodo}")

                                def extrair_ano(texto):
                                    match = re.search(r'\d{2}/\d{2}/(\d{4})', texto)
                                    if match:
                                        return int(match.group(1))
                                    return datetime.now().year
                                
                                ano_extrato = extrair_ano(texto_periodo)
                                logging.info(f"Ano extraído do período: {ano_extrato}")

                                meses_pt = {
                                    'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4, 'mai': 5, 'jun': 6,
                                    'jul': 7, 'ago': 8, 'set': 9, 'out': 10, 'nov': 11, 'dez': 12
                                }

                                def converter_data_pt(data_str, ano):
                                    try:
                                        dia, mes_abrev = data_str.strip().split('/')
                                        mes = meses_pt.get(mes_abrev.lower())
                                        if mes:
                                            data = datetime(ano, mes, int(dia))
                                            return data.strftime('%d/%m/%Y')
                                    except Exception as e:
                                        logging.info(f"Erro ao converter data '{data_str}': {e}")
                                    return data_str

                                logging.info("\nBuscando saldo inicial...")
                                saldo_inicial_celula = sheet['F9'].value
                                logging.info(f"Valor bruto encontrado em F9: {saldo_inicial_celula}")
                    
                                if isinstance(saldo_inicial_celula, str):
                                    saldo_inicial = float(saldo_inicial_celula.replace(".", "").replace(",", "."))
                                else:
                                    saldo_inicial = float(saldo_inicial_celula)
                        
                                saldo_inicial_frmt = locale.format_string("%.2f", saldo_inicial, grouping=True)
                                logging.info(f"Saldo inicial formatado: R${saldo_inicial_frmt}")
                    
                                resposta = messagebox.askyesno("Confirmação de saldo", f"O saldo inicial é de R${saldo_inicial_frmt}?")

                                if not resposta:
                                    logging.info("Usuário não confirmou o saldo inicial, solicitando entrada manual.")
                                    saldo_inicial_manual = simpledialog.askstring("Entrada de Saldo", "Insira o saldo inicial correto:")
                                    if saldo_inicial_manual:
                                        try:
                                            saldo_inicial = float(saldo_inicial_manual.replace(".", "").replace(",", "."))
                                            saldo_inicial_frmt = locale.format_string("%.2f", saldo_inicial, grouping=True)
                                        except ValueError:
                                            messagebox.showerror("Erro", "Valor de saldo inicial inválido.")
                                            return
                                    else:
                                        messagebox.showinfo("Aviso", "Processo cancelado pelo usuário.")
                                        return

                                self.saldo_inicial_entry.delete(0, tk.END)
                                self.saldo_inicial_entry.insert(0, saldo_inicial_frmt)
                    
                                saldo_final_calculado = saldo_inicial
                    
                                logging.info("\n=== INICIANDO PROCESSAMENTO DAS LINHAS ===")
                                for row in range(10, sheet.max_row + 1):
                                    try:
                                        logging.info(f"\nProcessando linha {row}:")
                                        data = sheet.cell(row=row, column=1).value
                                        logging.info(f"Data encontrada: {data}")
                            
                                        if not data:
                                            logging.info("Linha vazia, pulando...")
                                            continue
                                        if isinstance(data, str) and "total" in data.lower():
                                            logging.info("Encontrada linha de total, parando processamento...")
                                            break

                                        logging.info(f"Processando valores...")
                                        valor_celula = sheet.cell(row=row, column=6).value
                                        if valor_celula == "0,00":
                                            logging.info("Valor 0,00 encontrado, pulando linha...")
                                            continue

                                        logging.info(f"Processando descrições...")
                                        valor_celula = sheet.cell(row=row, column=2).value
                                        if valor_celula == "SALDO POUPANCA PLUS":
                                            logging.info("Saldo da poupança encontrado, pulando linha...")
                                            continue
                                
                                        historico = sheet.cell(row=row, column=2).value
                                        num_doc = sheet.cell(row=row, column=5).value
                                        valor = sheet.cell(row=row, column=6).value
                            
                                        logging.info(f"Valores lidos:")
                                        logging.info(f"  Histórico: {historico}, Nº Doc: {num_doc}, Valor: {valor}")
                            
                                        def converter_para_float(valor):
                                            if valor is None or valor == "":
                                                return 0.0
                                            if isinstance(valor, str):
                                                valor = valor.replace(".", "").replace(",", ".")
                                            try:
                                                return float(valor)
                                            except ValueError:
                                                return 0.0
                            
                                        valor_total = converter_para_float(valor)
                                        valor_formatado = formatar_valor_brasileiro(valor_total)
                                        logging.info(f"Valor total calculado: {valor_total}")

                                        if isinstance(data, datetime):
                                            data = data.strftime("%d/%m/%Y")
                                        elif isinstance(data, str):
                                            if re.match(r"\d{2}/[a-zA-Z]{3}", data):  #? ex: 02/jan
                                                data = converter_data_pt(data, ano_extrato)
                                            elif re.match(r"\d{2}/\d{2}", data):  #? ex: 02/01
                                                try:
                                                    data = datetime.strptime(f"{data}/{ano_extrato}", "%d/%m/%Y").strftime("%d/%m/%Y")
                                                except Exception as e:
                                                    logging.info(f"Erro ao converter data '{data}': {e}")
                            
                                        dados_importados.append([
                                            data, historico, num_doc, valor_formatado, "",
                                            "", "", "", "", "", "", "", "", ""
                                        ])
                            
                                        saldo_final_calculado += valor_formatado
                                        logging.info(f"Novo saldo calculado: {saldo_final_calculado}")
                            
                                    except Exception as e:
                                        logging.info(f"ERRO ao processar linha {row}:")
                                        logging.info(f"Detalhes do erro: {str(e)}")
                                        continue

                            logging.info("\n=== ATUALIZANDO INTERFACE ===")
                            logging.info("Formatando saldo final...")
                            saldo_final_calculado_frmt = locale.format_string("%.2f", saldo_final_calculado, grouping=True)
                            logging.info(f"Saldo final formatado: R${saldo_final_calculado_frmt}")
                
                            logging.info("Atualizando campo de saldo final...")
                            self.saldo_final_calculado_entry.delete(0, tk.END)
                            self.saldo_final_calculado_entry.insert(0, saldo_final_calculado_frmt)
                
                            logging.info("\nLimpando Treeview...")
                            for i in self.tree.get_children():
                                self.tree.delete(i)
                    
                            logging.info("Inserindo dados na Treeview...")
                            logging.info(f"Total de registros a inserir: {len(dados_importados)}")

                            for i, dados in enumerate(dados_importados):
                                tag = 'linha_par' if i % 2 == 0 else 'linha_impar'
                                self.tree.insert("", "end", values=dados, tags=(tag,))

                            self.atualizar_total_linhas_importadas()
                    
                            logging.info("\n=== PROCESSAMENTO CONCLUÍDO COM SUCESSO ===")
                            logging.info(f"Total de linhas processadas: {len(dados_importados)}")

                        except Exception as e:
                            logging.info("\n=== ERRO FATAL ===")
                            logging.info(f"Erro: {str(e)}")
                            logging.info("stack trace:")
                            traceback.print_exc()
                            messagebox.showerror("Erro",
                                "Erro ao processar o arquivo. Verifique se:\n\n" +
                                "1. O arquivo está no formato correto\n" +
                                "2. O arquivo não está em modo de exibição protegida\n" +
                                "3. O arquivo está fechado no Excel\n\n" +
                                f"Erro: {str(e)}")
                            return

                    else: #! LÓGICA PARA ARQUIVO XLS/XLSX, NOVO FORMATO E CONTA CORRENTE
                        logging.info("Processando XLS/XLSX, novo formato do Safra, conta corrente.")
                        import pdfplumber
                        import re
                        from tkinter import messagebox

                        try:
                            dados_importados = []
                            saldo_final_calculado = 0

                            if extensao == "xls": #! SE O ARQUIVO É XLS
                                logging.info("\n=== PROCESSANDO ARQUIVO XLS ===")
                                logging.info(f"Arquivo recebido: {arquivo}")
                                import xlrd
                                wb = xlrd.open_workbook(arquivo)
                                sheet = wb.sheet_by_index(0)
                                logging.info(f"Planilha aberta: {sheet.name}")

                                texto_periodo = sheet.cell_value(5, 0)
                                logging.info(f"Texto do período: {texto_periodo}")

                                def extrair_ano(texto):
                                    match = re.search(r'\d{2}/\d{2}/(\d{4})', texto)
                                    if match:
                                        return int(match.group(1))
                                    return datetime.now().year
                                
                                ano_extrato = extrair_ano(texto_periodo)
                                logging.info(f"Ano extraído do período: {ano_extrato}")

                                meses_pt = {
                                    'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4, 'mai': 5, 'jun': 6,
                                    'jul': 7, 'ago': 8, 'set': 9, 'out': 10, 'nov': 11, 'dez': 12
                                }

                                def converter_data_pt(data_str, ano):
                                    try:
                                        dia, mes_abrev = data_str.strip().split('/')
                                        mes = meses_pt.get(mes_abrev.lower())
                                        if mes:
                                            data = datetime(ano, mes, int(dia))
                                            return data.strftime('%d/%m/%Y')
                                    except Exception as e:
                                        logging.info(f"Erro ao converter data '{data_str}': {e}")
                                    return data_str

                                logging.info(f"Dimensões: {sheet.nrows} linhas x {sheet.ncols} colunas")

                                logging.info("\nBuscando saldo inicial...")
                                saldo_inicial = sheet.cell_value(12, 7)
                                logging.info(f"Valor bruto encontrado em F4: {saldo_inicial}")
                                logging.info(f"Tipo do valor: {type(saldo_inicial)}")

                                if isinstance(saldo_inicial, str):
                                    logging.info("Convertendo saldo inicial de string para float...")
                                    saldo_inicial = float(saldo_inicial.replace(".", "").replace(",", "."))
                                saldo_inicial_frmt = locale.format_string("%.2f", saldo_inicial, grouping=True)
                                logging.info(f"Saldo inicial formatado: R${saldo_inicial_frmt}")

                                resposta = messagebox.askyesno("Confirmação de saldo", f"O saldo inicial é de R${saldo_inicial_frmt}?")

                                if not resposta:
                                    logging.info("Usuário não confirmou o saldo inicial, solicitando entrada manual.")
                                    saldo_inicial_manual = simpledialog.askstring("Entrada de Saldo", "Insira o saldo inicial correto:")
                                    if saldo_inicial_manual:
                                        try:
                                            saldo_inicial = float(saldo_inicial_manual.replace(".", "").replace(",", "."))
                                            saldo_inicial_frmt = locale.format_string("%.2f", saldo_inicial, grouping=True)
                                        except ValueError:
                                            messagebox.showerror("Erro", "Valor de saldo inicial inválido.")
                                            return
                                    else:
                                        messagebox.showinfo("Aviso", "Processo cancelado pelo usuário.")
                                        return
                    
                                logging.info("Atualizando campo de saldo inicial na interface...")
                                self.saldo_inicial_entry.delete(0, tk.END)
                                self.saldo_inicial_entry.insert(0, saldo_inicial_frmt)

                                saldo_final_calculado = saldo_inicial

                                logging.info("\n=== INICIANDO PROCESSAMENTO DAS LINHAS ===")
                                logging.info(f"Total de linhas na planilha: {sheet.nrows}")

                                for row in range(13, sheet.nrows):
                                    try:
                                        logging.info(f"\nProcessando linha {row+1}:")
                                        data = sheet.cell_value(row, 0)
                                        logging.info(f"Data encontrada: {data} (tipo: {type(data)})")

                                        historico = sheet.cell_value(row, 3)
                                        num_doc = sheet.cell_value(row, 5)
                                        valor = sheet.cell_value(row, 6)

                                        logging.info(f"Valores lidos:")
                                        logging.info(f"  Histórico: {historico}, N° Doc: {num_doc}, Valor: {valor}")

                                        if valor is None or str(valor).strip() == "":
                                            logging.info("Valor vazio, pulando linha...")
                                            continue

                                        def converter_para_float(valor):
                                            logging.info(f"Tratando valor: {valor} (tipo: {type(valor)})")
                                            if valor is None or valor == "":
                                                logging.info("Valor vazio, retornando 0.0")
                                                return 0.0
                                            if isinstance(valor, str):
                                                logging.info("Convertendo string para float...")
                                                valor = valor.replace(".", "").replace(",", ".")
                                            try:
                                                resultado = float(valor)
                                                logging.info(f"Valor convertido: {resultado}")
                                                return resultado
                                            except ValueError as e:
                                                logging.info(f"Erro ao converter valor: {e}")
                                                return 0.0

                                        valor_total = converter_para_float(valor)
                                        valor_formatado = formatar_valor_brasileiro(valor_total)
                                        logging.info(f"Valor total calculado: {valor_formatado}")

                                        if isinstance(data, float):
                                            data = xlrd.xldate_as_datetime(data, wb.datemode).strftime('%d/%m/%Y')
                                        elif isinstance(data, str):
                                            if re.match(r'\d{2}/\d{2}', data):  #? ex: 02/01
                                                try:
                                                    data = datetime.strptime(f"{data}/{ano_extrato}", "%d/%m/%Y").strftime("%d/%m/%Y")
                                                except Exception as e:
                                                    logging.info(f"Erro ao converter data '{data}': {e}")
                                            elif re.match(r'\d{2}/[a-zA-Z]{3}', data):  #? ex: 02/jan
                                                data = converter_data_pt(data, ano_extrato)
                            
                                        logging.info("Adicionando linha aos dados importados...")
                                        dados_importados.append([
                                            data, historico, num_doc, valor_formatado, "",
                                            "", "", "", "", "", "", "", ""
                                        ])

                                        saldo_final_calculado += valor_total
                                        logging.info(f"Novo saldo calculado: {saldo_final_calculado}")
                            
                                    except Exception as e:
                                        logging.info(f"ERRO ao processar linha {row+1}:")
                                        logging.info(f"Detalhes do erro: {str(e)}")
                                        traceback.print_exc()
                                        continue

                            else: #! SE O ARQUIVO É XLSX
                                logging.info("\n=== PROCESSANDO ARQUIVO XLSX ===")
                                logging.info(f"Arquivo recebido: {arquivo}")
                                wb = openpyxl.load_workbook(arquivo, data_only=True)
                                sheet = wb.active
                                logging.info(f"Planilha ativa: {sheet.title}")

                                texto_periodo = sheet["A6"].value
                                logging.info(f"Texto do período: {texto_periodo}")

                                def extrair_ano(texto):
                                    match = re.search(r'\d{2}/\d{2}/(\d{4})', texto)
                                    if match:
                                        return int(match.group(1))
                                    return datetime.now().year
                                
                                ano_extrato = extrair_ano(texto_periodo)
                                logging.info(f"Ano extraído do período: {ano_extrato}")

                                meses_pt = {
                                    'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4, 'mai': 5, 'jun': 6,
                                    'jul': 7, 'ago': 8, 'set': 9, 'out': 10, 'nov': 11, 'dez': 12
                                }

                                def converter_data_pt(data_str, ano):
                                    try:
                                        dia, mes_abrev = data_str.strip().split('/')
                                        mes = meses_pt.get(mes_abrev.lower())
                                        if mes:
                                            data = datetime(ano, mes, int(dia))
                                            return data.strftime('%d/%m/%Y')
                                    except Exception as e:
                                        logging.info(f"Erro ao converter data '{data_str}': {e}")
                                    return data_str

                                logging.info("\nBuscando saldo inicial...")
                                saldo_inicial_celula = sheet['H13'].value
                                logging.info(f"Valor bruto encontrado em H13: {saldo_inicial_celula}")

                                if saldo_inicial_celula is None:
                                    messagebox.showerror("Erro", "A célula do saldo inicial está vazia.")
                                    return

                                if isinstance(saldo_inicial_celula, str):
                                    saldo_inicial = float(saldo_inicial_celula.replace(".", "").replace(",", "."))
                                else:
                                    saldo_inicial = float(saldo_inicial_celula)

                                saldo_inicial_frmt = locale.format_string("%.2f", saldo_inicial, grouping=True)
                                logging.info(f"Saldo inicial formatado: R${saldo_inicial_frmt}")
                    
                                resposta = messagebox.askyesno("Confirmação de saldo", f"O saldo inicial é de R${saldo_inicial_frmt}?")

                                if not resposta:
                                    logging.info("Usuário não confirmou o saldo inicial, solicitando entrada manual.")
                                    saldo_inicial_manual = simpledialog.askstring("Entrada de Saldo", "Insira o saldo inicial correto:")
                                    if saldo_inicial_manual:
                                        try:
                                            saldo_inicial = float(saldo_inicial_manual.replace(".", "").replace(",", "."))
                                            saldo_inicial_frmt = locale.format_string("%.2f", saldo_inicial, grouping=True)
                                        except ValueError:
                                            messagebox.showerror("Erro", "Valor de saldo inicial inválido.")
                                            return
                                    else:
                                        messagebox.showinfo("Aviso", "Processo cancelado pelo usuário.")
                                        return
                                    
                                self.saldo_inicial_entry.delete(0, tk.END)
                                self.saldo_inicial_entry.insert(0, saldo_inicial_frmt)

                                saldo_final_calculado = saldo_inicial

                                logging.info("\n=== INICIANDO PROCESSAMENTO DAS LINHAS ===")
                                for row in range(14, sheet.max_row + 1):
                                    try:
                                        logging.info(f"\nProcessando linha {row}:")
                                        data = sheet.cell(row=row, column=1).value
                                        logging.info(f"Data encontrada: {data}")

                                        valor = sheet.cell(row=row, column=7).value
                                        if not valor:
                                            logging.info("Linha vazia, pulando...")
                                        
                                        if isinstance(data, str) and "total" in data.lower():
                                            logging.info("Encontrada linha de total, parando processamento...")
                                            break

                                        logging.info(f"Processando valores...")
                                        valor_celula = sheet.cell(row=row, column=7).value
                                        if valor_celula == "0,00":
                                            logging.info("Valor 0,00 encontrado, pulando linha...")
                                            continue

                                        logging.info(f"Processando descrições...")
                                        valor_celula = sheet.cell(row=row, column=4).value
                                        if valor_celula == "SALDO CONTA CORRENTE":
                                            logging.info("Saldo da poupança encontrado, pulando linha...")
                                            continue

                                        historico = sheet.cell(row=row, column=4).value
                                        num_doc = sheet.cell(row=row, column=6).value
                                        valor = sheet.cell(row=row, column=7).value

                                        logging.info(f"Valores lidos:")
                                        logging.info(f"  Histórico: {historico}, Nº Doc: {num_doc}, Valor: {valor}")

                                        def converter_para_float(valor):
                                            if valor is None or valor == "":
                                                return 0.0
                                            if isinstance(valor, str):
                                                valor = valor.replace(".", "").replace(",", ".")
                                            try:
                                                return float(valor)
                                            except ValueError:
                                                return 0.0
                                            
                                        valor_total = converter_para_float(valor)
                                        valor_formatado = formatar_valor_brasileiro(valor_total)
                                        logging.info(f"Valor total calculado: {valor_total}")

                                        if isinstance(data, datetime):
                                            data = data.strftime("%d/%m/%Y")
                                        elif isinstance(data, str):
                                            if re.match(r"\d{2}/[a-zA-Z]{3}", data):  #? ex: 02/jan
                                                data = converter_data_pt(data, ano_extrato)
                                            elif re.match(r"\d{2}/\d{2}", data):  #? ex: 02/01
                                                try:
                                                    data = datetime.strptime(f"{data}/{ano_extrato}", "%d/%m/%Y").strftime("%d/%m/%Y")
                                                except Exception as e:
                                                    logging.info(f"Erro ao converter data '{data}': {e}")

                                        dados_importados.append([
                                            data, historico, num_doc, valor_formatado, "",
                                            "", "", "", "", "", "", "", "", ""
                                        ])

                                        saldo_final_calculado += valor_total
                                        logging.info(f"Novo saldo calculado: {saldo_final_calculado}")
                            
                                    except Exception as e:
                                        logging.info(f"ERRO ao processar linha {row}:")
                                        logging.info(f"Detalhes do erro: {str(e)}")
                                        continue

                            logging.info("\n=== ATUALIZANDO INTERFACE ===")
                            logging.info("Formatando saldo final...")
                            saldo_final_calculado_frmt = locale.format_string("%.2f", saldo_final_calculado, grouping=True)
                            logging.info(f"Saldo final formatado: R${saldo_final_calculado_frmt}")

                            logging.info("Atualizando campo de saldo final...")
                            self.saldo_final_calculado_entry.delete(0, tk.END)
                            self.saldo_final_calculado_entry.insert(0, saldo_final_calculado_frmt)
                
                            logging.info("\nLimpando Treeview...")
                            for i in self.tree.get_children():
                                self.tree.delete(i)
                    
                            logging.info("Inserindo dados na Treeview...")
                            logging.info(f"Total de registros a inserir: {len(dados_importados)}")

                            for i, dados in enumerate(dados_importados):
                                tag = 'linha_par' if i % 2 == 0 else 'linha_impar'

                                dados[3] = formatar_valor_brasileiro(dados[3])

                                self.tree.insert("", "end", values=dados, tags=(tag,))

                            self.atualizar_total_linhas_importadas()

                            logging.info("\n=== PROCESSAMENTO CONCLUÍDO COM SUCESSO ===")
                            logging.info(f"Total de linhas processadas: {len(dados_importados)}")

                        except Exception as e:
                            logging.info("\n=== ERRO FATAL ===")
                            logging.info(f"Erro: {str(e)}")
                            logging.info("stack trace:")
                            traceback.print_exc()
                            messagebox.showerror("Erro",
                                "Erro ao processar o arquivo. Verifique se:\n\n" +
                                "1. O arquivo está no formato correto\n" +
                                "2. O arquivo não está em modo de exibição protegida\n" +
                                "3. O arquivo está fechado no Excel\n\n" +
                                f"Erro: {str(e)}")
                            return

                else: #! LÓGICA PARA ARQUIVO XLS/XLSX, FORMATO ANTIGO
                    logging.info("Processando XLS/XLSX, formato antigo do Safra")

            elif extensao == "pdf":
                if novo_formato:
                    if conta_vinculada: #! ARQUIVO PDF, NOVO FORMATO E CONTA VINCULADA
                        logging.info("\n=== PROCESSANDO ARQUIVO PDF: CONTA VINCULADA ===")
                        logging.info(f"Arquivo Recebido: {arquivo}")
                        dados_importados = []
                        transacoes = []

                        import pdfplumber
                        import re
                        from tkinter import messagebox

                        try:
                            saldo_final_calculado = 0
                            linhas_processadas = []

                            with pdfplumber.open(arquivo) as pdf:
                                for pagina in pdf.pages:
                                    texto = pagina.extract_text()
                                    if texto:
                                        linhas_processadas.extend([linha.strip() for linha in texto.split('\n') if linha.strip()])

                            ano_extrato = None
                            for linha in linhas_processadas:
                                if "Período de" in linha:
                                    match = re.search(r'Período de \d{2}/\d{2}/(\d{4})', linha)
                                    if match:
                                        ano_extrato = match.group(1)
                                        break

                            logging.info(f"📄 Total de linhas extraídas do PDF: {len(linhas_processadas)}")

                            data_regex = re.compile(r'^\d{2}/\d{2}')

                            i = 0
                            while i < len(linhas_processadas):
                                linha = linhas_processadas[i].strip()

                                if "CONTA VINCULADA" in linha.upper() or "SALDO POUPANCA PLUS" in linha.upper():
                                    i += 1
                                    continue

                                if data_regex.match(linha):
                                    partes = linha.split()
                                    if len(partes) < 3:
                                        continue
                                    
                                    data = f"{partes[0]}/{ano_extrato}" if ano_extrato else partes[0]
                                    valor_str = partes[-1]
                                    documento = partes[-2] if partes[-2] != '-' else ''
                                    descricao = " ".join(partes[1:-2])

                                    if i + 1 < len(linhas_processadas):
                                        prox_linha = linhas_processadas[i + 1].strip()
                                        if not data_regex.match(prox_linha) and "CONTA VINCULADA" not in prox_linha.upper() and "SALDO POUPANCA PLUS" not in prox_linha.upper():
                                            descricao += " " + prox_linha
                                            i += 1

                                    valor = self.corrigir_valor(valor_str)
                                    credito = valor_str if valor > 0 else ""
                                    debito = valor_str if valor < 0 else ""

                                    transacao = [data, descricao, documento, credito or debito or "", "", "", "", "", "", "", "", "", "", ""]
                                    transacoes.append(transacao)
                                    dados_importados.append(transacao)
                                    saldo_final_calculado += valor
                                i += 1

                            logging.info(f"🔎 Total de transações agrupadas: {len(transacoes)}")

                            for i, t in enumerate(transacoes):
                                tag = 'linha_par' if i % 2 == 0 else 'linha_impar'
                                self.tree.insert("", "end", values=t, tags=(tag,))

                            self.atualizar_total_linhas_importadas()
                            saldo_final_frmt = locale.format_string("%.2f", saldo_final_calculado, grouping=True)
                            self.saldo_final_calculado_entry.delete(0, tk.END)
                            self.saldo_final_calculado_entry.insert(0, saldo_final_frmt)

                            logging.info(f"✅ {len(transacoes)} transações extraídas com sucesso!")
                            messagebox.showinfo("Sucesso", f"{len(transacoes)} transações importadas do PDF Safra!")

                        except Exception as e:
                            logging.info(f"❌ Erro ao processar PDF: {e}")
                            traceback.print_exc()
                            messagebox.showerror("Erro", f"Erro ao processar o PDF Safra:\n\n{str(e)}")

                    else: #! LÓGICA PARA ARQUIVO PDF, NOVO FORMATO E CONTA CORRENTE
                        logging.info("\n=== PROCESSANDO ARQUIVO PDF: CONTA CORRENTE ===")
                        logging.info(f"Arquivo Recebido: {arquivo}")
                        dados_importados = []
                        transacoes = []

                        try:
                            import pdfplumber
                            import re
                            from tkinter import messagebox

                            saldo_final_calculado = 0
                            linhas_processadas = []

                            with pdfplumber.open(arquivo) as pdf:
                                for pagina in pdf.pages:
                                    texto = pagina.extract_text()
                                    if texto:
                                        linhas = [linha.strip() for linha in texto.split('\n') if linha.strip()]
                                        linhas_processadas.extend(linhas)

                            ano_extrato = None
                            for linha in linhas_processadas:
                                if "Período de" in linha:
                                    match = re.search(r'Período de \d{2}/\d{2}/(\d{4})', linha)
                                    if match:
                                        ano_extrato = match.group(1)
                                        break

                            logging.info(f"📄 Total de linhas extraídas do PDF (Corrente): {len(linhas_processadas)}")

                            data_regex = re.compile(r'^\d{2}/\d{2}')
                            i = 0
                            saldo_final_definido = False

                            while i < len(linhas_processadas):
                                linha = linhas_processadas[i].strip()

                                if data_regex.match(linha):
                                    partes = linha.split()
                                    if len(partes) < 3:
                                        i += 1
                                        continue

                                    data = f"{partes[0]}/{ano_extrato}" if ano_extrato else partes[0]
                                    valor_str = partes[-1]
                                    documento = partes[-2] if partes[-2] != '-' else ''
                                    descricao = " ".join(partes[1:-2])

                                    if i + 1 < len(linhas_processadas):
                                        prox_linha = linhas_processadas[i + 1].strip()
                                        if not data_regex.match(prox_linha):
                                            descricao += " " + prox_linha
                                            i += 1

                                    if "SALDO CONTA" in descricao.upper():
                                        if not saldo_final_definido:
                                            valor = self.corrigir_valor(valor_str)
                                            saldo_final_frmt = locale.format_string("%.2f", valor, grouping=True)
                                            self.saldo_final_entry.delete(0, tk.END)
                                            self.saldo_final_entry.insert(0, saldo_final_frmt)
                                            saldo_final_definido = True
                                        i += 1
                                        continue

                                    valor = self.corrigir_valor(valor_str)
                                    credito = valor_str if valor > 0 else ""
                                    debito = valor_str if valor < 0 else "" 

                                    transacao = [data, descricao, documento, credito or debito or "", "", "", "", "", "", "", "", "", "", ""]
                                    transacoes.append(transacao)
                                    dados_importados.append(transacao)
                                    saldo_final_calculado += valor

                                    # Obter e converter o saldo final importado do campo de entrada
                                    saldo_final_importado_str = self.saldo_final_entry.get().replace(".", "").replace(",", ".")
                                    try:
                                        saldo_final_importado = float(saldo_final_importado_str)
                                    except ValueError:
                                        saldo_final_importado = 0.0

                                    # Calcular a diferença
                                    diferenca = saldo_final_importado - saldo_final_calculado

                                    # Exibir a diferença no campo apropriado
                                    diferenca_frmt = locale.format_string("%.2f", diferenca, grouping=True)
                                    self.diferenca_entry.delete(0, tk.END)
                                    self.diferenca_entry.insert(0, diferenca_frmt)

                                i += 1

                            logging.info(f"🔎 Total de transações conta corrente agrupadas: {len(transacoes)}")

                            for i, t in enumerate(transacoes):
                                tag = 'linha_par' if i % 2 == 0 else 'linha_impar'
                                self.tree.insert("", "end", values=t, tags=(tag,))

                            self.atualizar_total_linhas_importadas()
                            saldo_final_frmt = locale.format_string("%.2f", saldo_final_calculado, grouping=True)
                            self.saldo_final_calculado_entry.delete(0, tk.END)
                            self.saldo_final_calculado_entry.insert(0, saldo_final_frmt)

                            logging.info(f"✅ {len(transacoes)} transações de conta corrente importadas com sucesso!")
                            messagebox.showinfo("Sucesso", f"{len(transacoes)} transações importadas do PDF Safra (Conta Corrente)!")

                        except Exception as e:
                            logging.info(f"❌ Erro ao processar PDF: {e}")
                            traceback.print_exc()
                            messagebox.showerror("Erro", f"Erro ao processar o PDF Safra Conta Corrente:\n\n{str(e)}")

                else: #! LÓGICA PARA ARQUIVO PDF, FORMATO ANTIGO
                    logging.info("Processando PDF, formato antigo do Safra.")

            else:
                logging.info("Formato de arquivo não suportado.")
        else:
            logging.info("Nenhuma resposta específica fornecida para Safra.")

    def acao_santander(self, arquivo):
        logging.info("\n=== INÍCIO DO PROCESSAMENTO: SANTANDER ===")
        logging.info(f"Arquivo Recebido: {arquivo}")
        def formatar_valor_brasileiro(valor):
            try:
                return locale.format_string("%.2f", float(valor), grouping=True)
            except:
                return valor
        try:
            extensao = arquivo.lower().split('.')[-1]
            logging.info(f"Extensão detectada: {extensao}")
            dados_importados = []
            saldo_final_calculado = 0

            if extensao == "pdf": #! SE O ARQUIVO É PDF
                logging.info("\n=== PROCESSANDO ARQUIVO PDF ===")
                try:
                    import pdfplumber 
                    import re
                    with pdfplumber.open(arquivo) as pdf:

                        linhas = []
                        capturar = False

                        with pdfplumber.open(arquivo) as pdf:
                            for pagina in pdf.pages:
                                texto = pagina.extract_text()
                                if texto:
                                    for linha in texto.split("\n"):
                                        linha = linha.strip()
                                        if not linha:
                                            continue

                                        if not capturar and "SALDO ANTERIOR" in linha.upper():
                                            logging.info("💰 Detectado 'SALDO ANTERIOR'")
                                            match_valor = re.search(r"[-\d.,]+$", linha)
                                            if match_valor:
                                                saldo_str = match_valor.group()

                                                self.saldo_inicial_entry.delete(0, tk.END)
                                                self.saldo_inicial_entry.insert(0, saldo_str)

                                                try:
                                                    saldo_inicial_float = float(saldo_str.replace(".", "").replace(",", "."))
                                                except:
                                                    saldo_inicial_float = 0.0

                                                saldo_final_calculado = saldo_inicial_float

                                                logging.info(f"✅ Saldo inicial preenchido: {saldo_str}")
                                            capturar = True
                                            continue

                                        if capturar:
                                            linhas.append(linha)

                    logging.info(f"📄 Total de linhas extraídas do PDF: {len(linhas)}")

                    data_regex = re.compile(r'^\d{2}/\d{2}/\d{4}')
                    transacoes = []

                    for i, linha in enumerate(linhas):
                        logging.info(f"🔹 [{i}] Lendo linha: {linha}")
                        if not isinstance(linha, str) or not data_regex.match(linha):
                            continue

                        if "SALDO ANTERIOR" in linha.upper():
                            logging.info("🔍 Encontrado 'SALDO ANTERIOR'. Iniciando captura nas próximas linhas...")
                            continue

                        if not data_regex.match(linha):
                            continue

                        partes = linha.split()
                        data = partes[0]
                        doc_idx = next((i for i, p in enumerate(partes) if re.fullmatch(r'\d{6}', p)), -1)

                        if doc_idx == -1 or len(partes) < doc_idx + 2:
                            continue

                        documento = partes[doc_idx]
                        valor_str = partes[doc_idx + 1]
                        saldo_str = partes[doc_idx + 2] if len(partes) > doc_idx + 2 else ""
                        descricao = " ".join(partes[1:doc_idx])

                        valor_float = self.corrigir_valor(valor_str)
                        credito = valor_str if valor_float > 0 else ""
                        debito = valor_str if valor_float < 0 else ""

                        transacoes.append([
                            data, descricao, documento,
                            credito or debito or "",
                            saldo_str, "", "", "", "", "", "", "", "", ""
                        ])

                    logging.info(f"✅ {len(transacoes)} transações extraídas com sucesso!")

                    for i, t in enumerate(transacoes):
                        tag = 'linha_par' if i % 2 == 0 else 'linha_impar'
                        self.tree.insert("", "end", values=t, tags=(tag,))

                    self.atualizar_total_linhas_importadas()
                    saldo_final_frmt = locale.format_string("%.2f", saldo_final_calculado, grouping=True)
                    self.saldo_final_calculado_entry.delete(0, tk.END)
                    self.saldo_final_calculado_entry.insert(0, saldo_final_frmt)

                    messagebox.showinfo("Sucesso", f"{len(transacoes)} transações importadas do PDF Santander!")

                except Exception as e:
                    logging.info(f"❌ Erro ao processar PDF: {e}")
                    traceback.print_exc()
                    messagebox.showerror("Erro", f"Erro ao processar o PDF:\n\n{str(e)}")
                return

            elif extensao == 'xls': #! SE O ARQUIVO É XLS
                logging.info("\n=== PROCESSANDO ARQUIVO XLS ===")
                wb = xlrd.open_workbook(arquivo)
                sheet = wb.sheet_by_index(0)
                logging.info(f"Planilha aberta: {sheet.name}")
                logging.info(f"Dimensões? {sheet.nrows} linhas x {sheet.ncols} colunas")

                logging.info("\nBuscando saldo inicial...")
                saldo_inicial = sheet.cell_value(3, 5)
                logging.info(f"Valor bruto encontrado em F4: {saldo_inicial}")
                logging.info(f"Tipo do valor: {type(saldo_inicial)}")

                if isinstance(saldo_inicial, str):
                    logging.info("Convertendo saldo inicial de string para float...")
                    saldo_inicial = float(saldo_inicial.replace(".", "").replace(",", "."))
                saldo_inicial_frmt = locale.format_string("%.2f", saldo_inicial, grouping=True)
                logging.info(f"Saldo inicial formatado: R${saldo_inicial_frmt}")

                resposta = messagebox.askyesno("Confirmação de saldo", f"O saldo inicial é de R${saldo_inicial_frmt}?")

                if not resposta:
                    logging.info("Usuário não confirmou o saldo inicial, solicitando entrada manual.")
                    saldo_inicial_manual = simpledialog.askstring("Entrada de Saldo", "Insira o saldo inicial correto:")
                    if saldo_inicial_manual:
                        try:
                            saldo_inicial = float(saldo_inicial_manual.replace(".", "").replace(",", "."))
                            saldo_inicial_frmt = locale.format_string("%.2f", saldo_inicial, grouping=True)
                        except ValueError:
                            messagebox.showerror("Erro", "Valor de saldo inicial inválido.")
                            return
                    else:
                        messagebox.showinfo("Aviso", "Processo cancelado pelo usuário.")
                        return
                    
                logging.info("Atualizando campo de saldo inicial na interface...")
                self.saldo_inicial_entry.delete(0, tk.END)
                self.saldo_inicial_entry.insert(0, saldo_inicial_frmt)

                saldo_final_calculado = saldo_inicial

                logging.info("\n=== INICIANDO PROCESSAMENTO DAS LINHAS ===")
                logging.info(f"Total de linhas na planilha: {sheet.nrows}")

                for row in range(8, sheet.nrows):
                    try:
                        logging.info(f"\nProcessando linha {row+1}:")
                        data = sheet.cell_value(row, 0)
                        logging.info(f"Data encontrada: {data} (tipo: {type(data)})")

                        historico = sheet.cell_value(row, 2)
                        num_doc = sheet.cell_value(row, 3)
                        saldo = sheet.cell_value(row, 5)
                        valor = sheet.cell_value(row, 4)

                        logging.info(f"Valores lidos:")
                        logging.info(f"  Histórico: {historico}, N° Doc: {num_doc}, Valor: {valor}, Saldo: {saldo}")

                        if valor is None or str(valor).strip() == "":
                            logging.info("Valor vazio, pulando linha...")
                            continue

                        def converter_para_float(valor):
                            logging.info(f"Tratando valor: {valor} (tipo: {type(valor)})")
                            if valor is None or valor == "":
                                logging.info("Valor vazio, retornando 0.0")
                                return 0.0
                            if isinstance(valor, str):
                                logging.info("Convertendo string para float...")
                                valor = valor.replace(".", "").replace(",", ".")
                            try:
                                resultado = float(valor)
                                logging.info(f"Valor convertido: {resultado}")
                                return resultado
                            except ValueError as e:
                                logging.info(f"Erro ao converter valor: {e}")
                                return 0.0
                            
                        valor_total = converter_para_float(valor)
                        valor_formatado = formatar_valor_brasileiro(valor_total)
                        logging.info(f"Valor total calculado: {valor_formatado}")

                        saldo_total = converter_para_float(saldo)
                        saldo_formatado = formatar_valor_brasileiro(saldo_total)
                        logging.info(f"Saldo total calculado: {saldo_formatado}")

                        if isinstance(data, float):
                            logging.info("Convertendo data de float para string...")
                            data = xlrd.xldate_as_datetime(data, wb.datemode).strftime('%d/%m/%Y')
                            logging.info(f"Data convertida: {data}")
                            
                        logging.info("Adicionando linha aos dados importados...")
                        dados_importados.append([
                            data, historico, num_doc, valor_formatado, saldo_formatado,
                            "", "", "", "", "", "", "", ""
                        ])

                        saldo_final_calculado += valor_total
                        logging.info(f"Novo saldo calculado: {saldo_final_calculado}")
                            
                    except Exception as e:
                        logging.info(f"ERRO ao processar linha {row+1}:")
                        logging.info(f"Detalhes do erro: {str(e)}")
                        traceback.print_exc()
                        continue

            elif extensao == 'xlsx': #! SE O ARQUIVO É XLSX
                logging.info("\n=== PROCESSANDO ARQUIVO XLSX ===")
                wb = openpyxl.load_workbook(arquivo, data_only=True)
                sheet = wb.active
                logging.info(f"Planilha ativa: {sheet.title}")

                logging.info("\nBuscando saldo inicial...")
                saldo_inicial_celula = sheet['F4'].value
                logging.info(f"Valor bruto encontrado em F4: {saldo_inicial_celula}")

                if isinstance(saldo_inicial_celula, str):
                    saldo_inicial = float(saldo_inicial_celula.replace(".", "").replace(",", "."))
                else:
                    saldo_inicial = float(saldo_inicial_celula)

                saldo_inicial_frmt = locale.format_string("%.2f", saldo_inicial, grouping=True)
                logging.info(f"Saldo inicial formatado: R${saldo_inicial_frmt}")

                resposta = messagebox.askyesno("Confirmação de saldo", f"O saldo inicial é de R${saldo_inicial_frmt}?")

                if not resposta:
                    logging.info("Usuário não confirmou o saldo inicial, solicitando entrada manual.")
                    saldo_inicial_manual = simpledialog.askstring("Entrada de Saldo", "Insira o saldo inicial correto:")
                    if saldo_inicial_manual:
                        try:
                            saldo_inicial = float(saldo_inicial_manual.replace(".", "").replace(",", "."))
                            saldo_inicial_frmt = locale.format_string("%.2f", saldo_inicial, grouping=True)
                        except ValueError:
                            messagebox.showerror("Erro", "Valor de saldo inicial inválido.")
                            return
                    else:
                        messagebox.showinfo("Aviso", "Processo cancelado pelo usuário.")
                        return
                    
                self.saldo_inicial_entry.delete(0, tk.END)
                self.saldo_inicial_entry.insert(0, saldo_inicial_frmt)

                saldo_final_calculado = saldo_inicial

                logging.info("\n=== INICIANDO PROCESSAMENTO DAS LINHAS ===")
                for row in range(4, sheet.max_row + 1):
                    try:
                        logging.info(f"\nProcessando linha {row}:")
                        data = sheet.cell(row=row, column=1).value
                        logging.info(f"Data encontrada: {data}")

                        historico = sheet.cell(row=row, column=3).value
                        num_doc = sheet.cell(row=row, column=4).value
                        valor = sheet.cell(row=row, column=5).value
                        saldo = sheet.cell(row=row, column=6).value

                        logging.info(f"Valores lidos:")
                        logging.info(f"  Histórico: {historico}, N° Doc: {num_doc}, Valor: {valor}, Saldo: {saldo}")

                        def converter_para_float(valor):
                            if valor is None or valor == "":
                                return 0.0
                            if isinstance(valor, str):
                                valor = valor.replace(".", "").replace(",", ".")
                            try:
                                return float(valor)
                            except ValueError:
                                return 0.0
                            
                        valor_total = converter_para_float(valor)
                        valor_formatado = formatar_valor_brasileiro(valor_total)
                        logging.info(f"Valor total calculado: {valor_total}")

                        saldo_total = converter_para_float(saldo)
                        saldo_formatado = formatar_valor_brasileiro(saldo_total)
                        logging.info(f"Saldo total formatado: {saldo_total}")

                        dados_importados.append([
                            data, historico, num_doc, valor_formatado, saldo_formatado,
                            "", "", "", "", "", "", "", ""
                        ])

                        saldo_final_calculado += valor_total
                        logging.info(f"Novo saldo calculado: {saldo_final_calculado}")
                            
                    except Exception as e:
                        logging.info(f"ERRO ao processar linha {row}:")
                        logging.info(f"Detalhes do erro: {str(e)}")
                        continue

            logging.info("\n=== ATUALIZANDO INTERFACE ===")
            logging.info("Formatando saldo final...")
            saldo_final_calculado_frmt = locale.format_string("%.2f", saldo_final_calculado, grouping=True)
            logging.info(f"Saldo final formatado: R${saldo_final_calculado_frmt}")

            logging.info("Atualizando campo de saldo final...")
            self.saldo_final_calculado_entry.delete(0, tk.END)
            self.saldo_final_calculado_entry.insert(0, saldo_final_calculado_frmt)

            logging.info("\nLimpando Treeview...")
            for i in self.tree.get_children():
                self.tree.delete(i)

            logging.info("Inserindo dados na Treeview...")
            logging.info(f"Total de registros a inserir: {len(dados_importados)}")

            for i, dados in enumerate(dados_importados):
                tag = 'linha_par' if i % 2 == 0 else 'linha_impar'

                dados[3] = formatar_valor_brasileiro(dados[3])

                self.tree.insert("", "end", values=dados, tags=(tag,))

            self.atualizar_total_linhas_importadas()
                    
            logging.info("\n=== PROCESSAMENTO CONCLUÍDO COM SUCESSO ===")
            logging.info(f"Total de linhas processadas: {len(dados_importados)}")

        except Exception as e:
            logging.info("\n=== ERRO FATAL ===")
            logging.info(f"Erro: {str(e)}")
            logging.info("stack trace:")
            traceback.print_exc()
            messagebox.showerror("Erro",
                "Erro ao processar o arquivo. Verifique se:\n\n" +
                "1. O arquivo está no formato correto\n" +
                "2. O arquivo não está em modo de exibição protegida\n" +
                "3. O arquivo está fechado no Excel\n\n" +
                f"Erro: {str(e)}")
            return

    def acao_daycoval(self, arquivo):
        logging.info("\n=== INÍCIO DO PROCESSAMENTO: DAYCOVAL ===")
        logging.info(f"Arquivo Recebido: {arquivo}")
        try:
            extensao = arquivo.lower().split('.')[-1]
            logging.info(f"Extensão detectada: {extensao}")
            dados_importados = []
            saldo_final_calculado = 0
            def formatar_valor_brasileiro(valor):
                try:
                    return locale.format_string("%.2f", float(valor), grouping=True)
                except:
                    return valor

            if extensao == 'xls': #! SE O ARQUIVO É XLS
                logging.info("\n=== PROCESSANDO ARQUIVO XLS ===")
                wb = xlrd.open_workbook(arquivo)
                sheet = wb.sheet_by_index(0)
                logging.info(f"Planilha aberta: {sheet.name}")

                meses_pt = {
                    'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4, 'mai': 5, 'jun': 6,
                    'jul': 7, 'ago': 8, 'set': 9, 'out': 10, 'nov': 11, 'dez': 12
                }

                def converter_data_pt(data_str, ano):
                    try:
                        dia, mes_abrev = data_str.strip().split('/')
                        mes = meses_pt.get(mes_abrev.lower())
                        if mes:
                            data = datetime(ano, mes, int(dia))
                            return data.strftime('%d/%m/%Y')
                    except Exception as e:
                        logging.info(f"Erro ao converter data '{data_str}': {e}")
                        return data_str
                    
                texto_periodo = sheet.cell_value(5, 0)
                logging.info(f"Texto do período: {texto_periodo}")
                def extrair_ano(texto):
                    match = re.search(r'\d{2}/\d{2}/(\d{4})', texto)
                    if match:
                        return datetime.now().year
                    ano_extrato = extrair_ano(texto_periodo)
                    logging.info(f"Ano extraído do período: {ano_extrato}")

                logging.info(f"Dimensões: {sheet.nrows} linhas x {sheet.ncols} colunas")

                logging.info("\nBuscando saldo inicial...")
                saldo_inicial = sheet.cell_value(9, 5)
                logging.info(f"Valor bruto encontrado em F10: {saldo_inicial}")
                logging.info(f"Tipo do valor: {type(saldo_inicial)}")

                if isinstance(saldo_inicial, str):
                    logging.info("Convertendo saldo inicial de string para float...")
                    saldo_inicial = float(saldo_inicial.replace(".", "").replace(",", "."))
                saldo_inicial_frmt = locale.format_string("%.2f", saldo_inicial, grouping=True)
                logging.info(f"Saldo inicial formatado: R${saldo_inicial_frmt}")

                resposta = messagebox.askyesno("Confirmação de saldo", f"O saldo inicial é de R${saldo_inicial_frmt}?")

                if not resposta:
                    logging.info("Usuário não confirmou o saldo inicial, solicitando entrada manual.")
                    saldo_inicial_manual = simpledialog.askstring("Entrada de Saldo", "Insira o saldo inicial correto:")
                    if saldo_inicial_manual:
                        try:
                            saldo_inicial = float(saldo_inicial_manual.replace(".", "").replace(",", "."))
                            saldo_inicial_frmt = locale.format_string("%.2f", saldo_inicial, grouping=True)
                        except ValueError:
                            messagebox.showerror("Erro", "Valor de saldo inicial inválido.")
                            return
                    else:
                        messagebox.showinfo("Aviso", "Processo cancelado pelo usuário.")
                        return
                    
                logging.info("Atualizando campo de saldo inicial na interface...")
                self.saldo_inicial_entry.delete(0, tk.END)
                self.saldo_inicial_entry.insert(0, saldo_inicial_frmt)

                saldo_final_calculado = saldo_inicial

                logging.info("\n=== INICIANDO PROCESSAMENTO DAS LINHAS ===")
                logging.info(f"Total de linhas na planilha: {sheet.nrows}")

                for row in range(10, sheet.nrows):
                    try:
                        logging.info(f"\nProcessando linha {row+1}:")
                        data = sheet.cell_value(row, 0)
                        logging.info(f"Data encontrada: {data} (tipo: {type(data)})")

                        historico = sheet.cell_value(row, 2)
                        num_doc = sheet.cell_value(row, 1)
                        debito = sheet.cell_value(row, 3)
                        credito = sheet.cell_value(row, 4)
                        saldo = sheet.cell_value(row, 5)
                        valor = sheet.cell_value(row, 4)

                        logging.info(f"Valores lidos:")
                        logging.info(f"  Histórico: {historico}, N° Doc: {num_doc}, Débito: {debito}, Crédito: {credito}, Valor: {valor}, Saldo: {saldo}")

                        if valor is None or str(valor).strip() == "":
                            logging.info("Valor vazio, pulando linha...")
                            continue

                        def converter_para_float(valor):
                            logging.info(f"Tratando valor: {valor} (tipo: {type(valor)})")
                            if valor is None or valor == "":
                                logging.info("Valor vazio, retornando 0.0")
                                return 0.0
                            if isinstance(valor, str):
                                logging.info("Convertendo string para float...")
                                valor = valor.replace(".", "").replace(",", ".")
                            try:
                                resultado = float(valor)
                                logging.info(f"Valor convertido: {resultado}")
                                return resultado
                            except ValueError as e:
                                logging.info(f"Erro ao converter valor: {e}")
                                return 0.0
                            
                        valor_credito = converter_para_float(credito)
                        valor_debito = converter_para_float(debito)
                        valor_total = valor_credito + valor_debito
                        valor_formatado = formatar_valor_brasileiro(valor_total)
                        logging.info(f"Valor total calculado: {valor_formatado}")

                        saldo_total = converter_para_float(saldo)
                        saldo_formatado = formatar_valor_brasileiro(saldo_total)
                        logging.info(f"Saldo total calculado: {saldo_formatado}")

                        if isinstance(data, float):
                            logging.info("Convertendo data float padrão Excel.")
                            data = xlrd.xldate_as_datetime(data, wb.datemode).strftime('%d/%m/%Y')
                            logging.info(f"Data convertido: {data}")
                        elif isinstance(data, str):
                            logging.info(f"Convertendo data string com mês pt-br: {data}")
                            data = converter_data_pt(data, ano_extrato)
                            logging.info(f"Data convertida: {data}")
                            
                        logging.info("Adicionando linha aos dados importados...")
                        dados_importados.append([
                            data, historico, num_doc, valor_formatado, saldo_formatado,
                            "", "", "", "", "", "", "", ""
                        ])

                        saldo_final_calculado += valor_total
                        logging.info(f"Novo saldo calculado: {saldo_final_calculado}")
                            
                    except Exception as e:
                        logging.info(f"ERRO ao processar linha {row+1}:")
                        logging.info(f"Detalhes do erro: {str(e)}")
                        traceback.print_exc()
                        continue

            else:  #! SE O ARQUIVO É XLSX
                logging.info("\n=== PROCESSANDO ARQUIVO XLSX ===")
                wb = openpyxl.load_workbook(arquivo, data_only=True)
                sheet = wb.active
                logging.info(f"Planilha ativa: {sheet.title}")

                meses_pt = {
                    'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4, 'mai': 5, 'jun': 6,
                    'jul': 7, 'ago': 8, 'set': 9, 'out': 10, 'nov': 11, 'dez': 12
                }

                def converter_data_pt(data_str, ano):
                    try:
                        dia, mes_abrev = data_str.strip().split('/')
                        mes = meses_pt.get(mes_abrev.lower())
                        if mes:
                            data = datetime(int(ano), mes, int(dia))
                            return data.strftime('%d/%m/%Y')
                    except Exception as e:
                        logging.info(f"Erro ao converter data '{data_str}': {e}")
                    return data_str
                
                texto_periodo = sheet["A6"].value
                logging.info(f"Texto do período: {texto_periodo}")
                def extrair_ano(texto):
                    match = re.search(r'\d{2}/\d{2}/(\d{4})', texto)
                    if match:
                        return int(match.group(1))
                    return datetime.now().year
                ano_extrato = extrair_ano(texto_periodo)
                logging.info(f"Ano extraído do período: {ano_extrato}")

                logging.info("\nBuscando saldo inicial...")
                saldo_inicial_celula = sheet['F10'].value
                logging.info(f"Valor bruto encontrado em F10: {saldo_inicial_celula}")

                if isinstance(saldo_inicial_celula, str):
                    saldo_inicial = float(saldo_inicial_celula.replace(".", "").replace(",", "."))
                else:
                    saldo_inicial = float(saldo_inicial_celula)

                saldo_inicial_frmt = locale.format_string("%.2f", saldo_inicial, grouping=True)
                logging.info(f"Saldo inicial formatado: R${saldo_inicial_frmt}")

                resposta = messagebox.askyesno("Confirmação de saldo", f"O saldo inicial é de R${saldo_inicial_frmt}?")

                if not resposta:
                    logging.info("Usuário não confirmou o saldo inicial, solicitando entrada manual.")
                    saldo_inicial_manual = simpledialog.askstring("Entrada de Saldo", "Insira o saldo inicial correto:")
                    if saldo_inicial_manual:
                        try:
                            saldo_inicial = float(saldo_inicial_manual.replace(".", "").replace(",", "."))
                            saldo_inicial_frmt = locale.format_string("%.2f", saldo_inicial, grouping=True)
                        except ValueError:
                            messagebox.showerror("Erro", "Valor de saldo inicial inválido.")
                            return
                    else:
                        messagebox.showinfo("Aviso", "Processo cancelado pelo usuário.")
                        return
                    
                self.saldo_inicial_entry.delete(0, tk.END)
                self.saldo_inicial_entry.insert(0, saldo_inicial_frmt)

                saldo_final_calculado = saldo_inicial

                logging.info("\n=== INICIANDO PROCESSAMENTO DAS LINHAS ===")
                for row in range(11, sheet.max_row + 1):
                    try:
                        logging.info(f"\nProcessando linha {row}:")
                        data = sheet.cell(row=row, column=1).value
                        logging.info(f"Data encontrada: {data}")

                        historico = sheet.cell(row=row, column=3).value
                        num_doc = sheet.cell(row=row, column=2).value
                        debito = sheet.cell(row=row, column=4).value
                        credito = sheet.cell(row=row, column=5).value
                        valor = sheet.cell(row=row, column=5).value
                        saldo = sheet.cell(row=row, column=6).value

                        logging.info(f"Valores lidos:")
                        logging.info(f"  Histórico: {historico}, N° Doc: {num_doc}, Débito: {debito}, Crédito: {credito}, Valor: {valor}, Saldo: {saldo}")

                        def converter_para_float(valor):
                            if valor is None or valor == "":
                                return 0.0
                            if isinstance(valor, str):
                                valor = valor.replace(".", "").replace(",", ".")
                            try:
                                return float(valor)
                            except ValueError:
                                return 0.0
                            
                        valor_credito = converter_para_float(credito)
                        valor_debito = converter_para_float(debito)
                        valor_total = valor_credito + valor_debito
                        valor_formatado = formatar_valor_brasileiro(valor_total)
                        logging.info(f"Valor total calculado: {valor_formatado}")

                        saldo_total = converter_para_float(saldo)
                        saldo_formatado = formatar_valor_brasileiro(saldo_total)
                        logging.info(f"Saldo total formatado: {saldo_total}")

                        if isinstance(data, datetime):
                            logging.info("Data já está no formato datetime...")
                            data = data.strftime("%d/%m/%Y")
                        elif isinstance(data, str):
                            logging.info(f"Convertendo data abreviada em pt-br: {data}")
                            data = converter_data_pt(data, ano_extrato)
                            logging.info(f"Data convertido: {data}")

                        dados_importados.append([
                            data, historico, num_doc, valor_formatado, saldo_formatado,
                            "", "", "", "", "", "", "", ""
                        ])

                        saldo_final_calculado += valor_total
                        logging.info(f"Novo saldo calculado: {saldo_final_calculado}")
                            
                    except Exception as e:
                        logging.info(f"ERRO ao processar linha {row}:")
                        logging.info(f"Detalhes do erro: {str(e)}")
                        continue

            logging.info("\n=== ATUALIZANDO INTERFACE ===")
            logging.info("Formatando saldo final...")
            saldo_final_calculado_frmt = locale.format_string("%.2f", saldo_final_calculado, grouping=True)
            logging.info(f"Saldo final formatado: R${saldo_final_calculado_frmt}")

            logging.info("Atualizando campo de saldo final...")
            self.saldo_final_calculado_entry.delete(0, tk.END)
            self.saldo_final_calculado_entry.insert(0, saldo_final_calculado_frmt)

            logging.info("\nLimpando Treeview...")
            for i in self.tree.get_children():
                self.tree.delete(i)

            logging.info("Inserindo dados na Treeview...")
            logging.info(f"Total de registros a inserir: {len(dados_importados)}")

            for i, dados in enumerate(dados_importados):
                tag = 'linha_par' if i % 2 == 0 else 'linha_impar'

                dados[3] = formatar_valor_brasileiro(dados[3])

                self.tree.insert("", "end", values=dados, tags=(tag,))

            self.atualizar_total_linhas_importadas()

            logging.info("\n=== PROCESSAMENTO CONCLUÍDO COM SUCESSO ===")
            logging.info(f"Total de linhas processadas: {len(dados_importados)}")

        except Exception as e:
            logging.info("\n=== ERRO FATAL ===")
            logging.info(f"Erro: {str(e)}")
            logging.info("Stack trace:")
            traceback.print_exc()
            messagebox.showerror("Erro", 
                "Erro ao processar o arquivo. Verifique se:\n\n" +
                "1. O arquivo está no formato correto\n" +
                "2. O arquivo não está em modo de exibição protegida\n" +
                "3. O arquivo está fechado no Excel\n\n" +
                f"Erro: {str(e)}")
            return

    def acao_itau(self, arquivo):
        logging.info("\n=== INÍCIO DO PROCESSAMENTO: ITAÚ ===")
        logging.info(f"Arquivo Recebido: {arquivo}")
        def formatar_valor_brasileiro(valor):
            try:
                return locale.format_string("%.2f", float(valor), grouping=True)
            except:
                return valor
        try:
            extensao = arquivo.lower().split('.')[-1]
            logging.info(f"Extensão detectada: {extensao}")
            dados_importados = []
            saldo_final_calculado = 0

            if extensao == "pdf": #! SE O ARQUIVO É PDF
                logging.info("\n=== PROCESSANDO ARQUIVO PDF ===")
                try:
                    import pdfplumber 
                    import re
                    with pdfplumber.open(arquivo) as pdf:

                        linhas = []
                        capturar = False

                        for pagina in pdf.pages:
                            texto = pagina.extract_text()
                            if texto:
                                for linha in texto.split("\n"):
                                    linha = linha.strip()
                                    if not linha:
                                        continue

                                    if not capturar and "SALDO ANTERIOR" in linha.upper():
                                        logging.info("💰 Detectado 'SALDO ANTERIOR'")
                                        match_valor = re.search(r"[-\d.,]+$", linha)
                                        if match_valor:
                                            saldo_str = match_valor.group()

                                            self.saldo_inicial_entry.delete(0, tk.END)
                                            self.saldo_inicial_entry.insert(0, saldo_str)

                                            try:
                                                saldo_inicial_float = float(saldo_str.replace(".", "").replace(",", "."))
                                            except:
                                                saldo_inicial_float = 0.0

                                            saldo_final_calculado = saldo_inicial_float

                                            logging.info(f"✅ Saldo inicial preenchido: {saldo_str}")
                                        capturar = True
                                        continue

                                    if capturar:
                                        linhas.append(linha)

                    logging.info(f"📄 Total de linhas extraídas do PDF: {len(linhas)}")

                    data_regex = re.compile(r'^\d{2}/\d{2}')
                    transacoes = []
                    self.tree.delete(*self.tree.get_children())

                    for i, linha in enumerate(linhas):
                        logging.info(f"🔹 [{i}] Lendo linha: {linha}")
                        if not isinstance(linha, str) or not data_regex.match(linha):
                            continue

                        partes = linha.split()
                        if len(partes) < 3:
                            continue
                        
                        data = partes[0]

                        valores_numericos = [p for p in partes if re.fullmatch(r"-?\d{1,3}(\.\d{3})*,\d{2}", p)]
                        if not valores_numericos:
                            logging.info(f"⏭️ Pulando linha sem valor identificado: {linha}")
                            continue

                        valor_str = valores_numericos[-1]
                        saldo_str = valores_numericos[-2] if len(valores_numericos) >= 2 else ""

                        try:
                            idx_valor = len(partes) - list(reversed(partes)).index(valor_str) - 1
                        except ValueError:
                            logging.info(f"⚠️ Valor não encontrado na linha: {linha}")
                            continue

                        if idx_valor == 2 and len(partes) == 3:
                            descricao = partes[1]
                        else:
                            descricao = " ".join(partes[1:idx_valor])
                        logging.info(f"🧾 Descrição final: '{descricao}'")

                        descricao_limpa = descricao.replace(" ", "").upper()
                        if descricao_limpa.startswith("SALD"):
                            logging.info(f"⏭️  Ignorando linha com descrição suspeita de saldo: {descricao}")
                            continue

                        valor_float = self.corrigir_valor(valor_str)
                        credito = valor_str if valor_float > 0 else ""
                        debito = valor_str if valor_float < 0 else ""

                        logging.info(f"📌 Transação válida detectada: {data} | {descricao} | {valor_str}")
                        transacoes.append([
                            data, descricao, "",  # documento em branco
                            credito or debito or "",
                            saldo_str, "", "", "", "", "", "", "", "", ""
                        ])

                    logging.info(f"✅ {len(transacoes)} transações extraídas com sucesso!")

                    for i, t in enumerate(transacoes):
                        tag = 'linha_par' if i % 2 == 0 else 'linha_impar'
                        self.tree.insert("", "end", values=t, tags=(tag,))

                    self.atualizar_total_linhas_importadas()
                    saldo_final_frmt = locale.format_string("%.2f", saldo_final_calculado, grouping=True)
                    self.saldo_final_calculado_entry.delete(0, tk.END)
                    self.saldo_final_calculado_entry.insert(0, saldo_final_frmt)

                    messagebox.showinfo("Sucesso", f"{len(transacoes)} transações importadas do PDF Itaú!")

                except Exception as e:
                    logging.info(f"❌ Erro ao processar PDF: {e}")
                    traceback.print_exc()
                    messagebox.showerror("Erro", f"Erro ao processar o PDF:\n\n{str(e)}")
                return

            elif extensao == "xls": #! SE O ARQUIVO É XLS
                logging.info("\n=== PROCESSANDO ARQUIVO XLS ===")
                wb = xlrd.open_workbook(arquivo)
                sheet = wb.sheet_by_index(0)
                logging.info(f"Planilha aberta: {sheet.name}")
                logging.info(f"Dimensões? {sheet.nrows} linhas x {sheet.ncols} colunas")

                logging.info("\nBuscando saldo inicial...")
                saldo_inicial = sheet.cell_value(7, 7)
                logging.info(f"Valor bruto encontrado em H8: {saldo_inicial}")
                logging.info(f"Tipo do valor: {type(saldo_inicial)}")

                if isinstance(saldo_inicial, str):
                    logging.info("Convertendo saldo inicial de string para float...")
                    saldo_inicial = float(saldo_inicial.replace(".", "").replace(",", "."))
                saldo_inicial_frmt = locale.format_string("%.2f", saldo_inicial, grouping=True)
                logging.info(f"Saldo inicial formatado: R${saldo_inicial_frmt}")

                resposta = messagebox.askyesno("Confirmação de saldo", f"O saldo inicial é de R${saldo_inicial_frmt}?")

                if not resposta:
                    logging.info("Usuário não confirmou o saldo inicial, solicitando entrada manual.")
                    saldo_inicial_manual = simpledialog.askstring("Entrada de Saldo", "Insira o saldo inicial correto:")
                    if saldo_inicial_manual:
                        try:
                            saldo_inicial = float(saldo_inicial_manual.replace(".", "").replace(",", "."))
                            saldo_inicial_frmt = locale.format_string("%.2f", saldo_inicial, grouping=True)
                        except ValueError:
                                messagebox.showerror("Erro", "Valor de saldo inicial inválido.")
                                return
                    else:
                        messagebox.showinfo("Aviso", "Processo cancelado pelo usuário.")
                        return
                    
                logging.info("Atualizando campo de saldo inicial na interface...")
                self.saldo_inicial_entry.delete(0, tk.END)
                self.saldo_inicial_entry.insert(0, saldo_inicial_frmt)

                saldo_final_calculado = saldo_inicial

                logging.info("\n=== INICIANDO PROCESSAMENTO DAS LINHAS ===")
                logging.info(f"Total de linhas na planilha: {sheet.nrows}")

                for row in range(8, sheet.nrows):
                    try:
                        logging.info(f"\nProcessando linha {row+1}:")
                        data = sheet.cell_value(row, 1)
                        logging.info(f"Data encontrada: {data} (tipo: {type(data)})")

                        historico = sheet.cell_value(row, 4)
                        valor = sheet.cell_value(row, 6)

                        logging.info(f"Valores lidos:")
                        logging.info(f"  Histórico: {historico}, Valor: {valor}")

                        if valor is None or str(valor).strip() == "":
                            logging.info("Valor vazio, pulando linha...")
                            continue

                        def converter_para_float(valor):
                            logging.info(f"Tratando valor: {valor} (tipo: {type(valor)})")
                            if valor is None or valor == "":
                                logging.info("Valor vazio, retornando 0.0")
                                return 0.0
                            if isinstance(valor, str):
                                logging.info("Convertendo string para float...")
                                valor = valor.replace(".", "").replace(",", ".")
                            try:
                                resultado = float(valor)
                                logging.info(f"Valor convertido: {resultado}")
                                return resultado
                            except ValueError as e:
                                logging.info(f"Erro ao converter valor: {e}")
                                return 0.0
                        
                        valor_total = valor
                        logging.info(f"Valor total calculado: {valor_total}")

                        if isinstance(data, float):
                            logging.info("Convertendo data de float para string...")
                            data = xlrd.xldate_as_datetime(data, wb.datemode).strftime('%d/%m/%Y')
                            logging.info(f"Data convertida: {data}")
                            
                        logging.info("Adicionando linha aos dados importados...")
                        dados_importados.append([
                            data, historico, "", valor_total, valor,
                            "", "", "", "", "", "", "", ""
                        ])

                        saldo_final_calculado += valor_total
                        logging.info(f"Novo saldo calculado: {saldo_final_calculado}")
                            
                    except Exception as e:
                        logging.info(f"ERRO ao processar linha {row+1}:")
                        logging.info(f"Detalhes do erro: {str(e)}")
                        traceback.print_exc()
                        continue

            elif extensao == "xlsx": #! SE O ARQUIVO É XLSX
                logging.info("\n=== PROCESSANDO ARQUIVO XLSX ===")
                wb = openpyxl.load_workbook(arquivo, data_only=True)
                sheet = wb.active
                logging.info(f"Planilha ativa: {sheet.title}")

                logging.info("\nBuscando saldo inicial...")
                saldo_inicial_celula = sheet['H8'].value
                logging.info(f"Valor bruto encontrado em H8: {saldo_inicial_celula}")

                if isinstance(saldo_inicial_celula, str):
                    saldo_inicial = float(saldo_inicial_celula.replace(".", "").replace(",", "."))
                else:
                    saldo_inicial = float(saldo_inicial_celula)

                saldo_inicial_frmt = locale.format_string("%.2f", saldo_inicial, grouping=True)
                logging.info(f"Saldo inicial formatado: R${saldo_inicial_frmt}")

                resposta = messagebox.askyesno("Confirmação de saldo", f"O saldo inicial é de R${saldo_inicial_frmt}?")

                if not resposta:
                    logging.info("Usuário não confirmou o saldo inicial, solicitando entrada manual.")
                    saldo_inicial_manual = simpledialog.askstring("Entrada de Saldo", "Insira o saldo inicial correto:")
                    if saldo_inicial_manual:
                        try:
                            saldo_inicial = float(saldo_inicial_manual.replace(".", "").replace(",", "."))
                            saldo_inicial_frmt = locale.format_string("%.2f", saldo_inicial, grouping=True)
                        except ValueError:
                            messagebox.showerror("Erro", "Valor de saldo inicial inválido.")
                            return
                    else:
                        messagebox.showinfo("Aviso", "Processo cancelado pelo usuário.")
                        return
                    
                self.saldo_inicial_entry.delete(0, tk.END)
                self.saldo_inicial_entry.insert(0, saldo_inicial_frmt)

                saldo_final_calculado = saldo_inicial

                logging.info("\n=== INICIANDO PROCESSAMENTO DAS LINHAS ===")
                for row in range(9, sheet.max_row + 1):
                    try:
                        logging.info(f"\nProcessando linha {row}:")
                        data = sheet.cell(row=row, column=2).value

                        if isinstance(data, datetime):
                            data = data.date()

                        elif not isinstance(data, date):
                            logging.info("Data inválida ou vazia, pulando...")
                            continue

                        data_formatada = data.strftime("%d/%m/%Y")
                        logging.info(f"Data formatada encontrada: {data_formatada}")

                        historico = sheet.cell(row=row, column=5).value
                        valor = sheet.cell(row=row, column=7).value

                        logging.info(f"Valores lidos:")
                        logging.info(f"  Histórico: {historico}, Valor: {valor}")

                        if valor is None or str(valor).strip() == "":
                            logging.info("Valor vazio, pulando linha...")
                            continue

                        def converter_para_float(valor):
                            if valor is None or valor == "":
                                return 0.0
                            if isinstance(valor, str):
                                valor = valor.replace(".", "").replace(",", ".")
                            try:
                                return float(valor)
                            except ValueError:
                                return 0.0
                            
                        valor_total = converter_para_float(valor)
                        valor_formatado = formatar_valor_brasileiro(valor_total)
                        logging.info(f"Valor total calculado: {valor_total}")
                            
                        dados_importados.append([
                            data_formatada, historico, "", valor_formatado, "",
                            "", "", "", "", "", "", "", ""
                        ])

                        saldo_final_calculado += valor_total
                        logging.info(f"Novo saldo calculado: {saldo_final_calculado}")
                            
                    except Exception as e:
                        logging.info(f"ERRO ao processar linha {row}:")
                        logging.info(f"Detalhes do erro: {str(e)}")
                        continue

            logging.info("\n=== ATUALIZANDO INTERFACE ===")
            logging.info("Formatando saldo final...")
            saldo_final_calculado_frmt = locale.format_string("%.2f", saldo_final_calculado, grouping=True)
            logging.info(f"Saldo final formatado: R${saldo_final_calculado_frmt}")

            logging.info("Atualizando campo de saldo final...")
            self.saldo_final_calculado_entry.delete(0, tk.END)
            self.saldo_final_calculado_entry.insert(0, saldo_final_calculado_frmt)

            logging.info("\nLimpando Treeview...")
            for i in self.tree.get_children():
                self.tree.delete(i)
                    
            logging.info("Inserindo dados na Treeview...")
            logging.info(f"Total de registros a inserir: {len(dados_importados)}")

            for i, dados in enumerate(dados_importados):
                tag = 'linha_par' if i % 2 == 0 else 'linha_impar'

                dados[3] = formatar_valor_brasileiro(dados[3])

                self.tree.insert("", "end", values=dados, tags=(tag,))

            self.atualizar_total_linhas_importadas()
                    
            logging.info("\n=== PROCESSAMENTO CONCLUÍDO COM SUCESSO ===")
            logging.info(f"Total de linhas processadas: {len(dados_importados)}")

        except Exception as e:
            logging.info("\n=== ERRO FATAL ===")
            logging.info(f"Erro: {str(e)}")
            logging.info("stack trace:")
            traceback.print_exc()
            messagebox.showerror("Erro",
                "Erro ao processar o arquivo. Verifique se:\n\n" +
                "1. O arquivo está no formato correto\n" +
                "2. O arquivo não está em modo de exibição protegida\n" +
                "3. O arquivo está fechado no Excel\n\n" +
                f"Erro: {str(e)}")
            return

    def acao_brasil(self, arquivo):
        logging.info("\n=== INÍCIO DO PROCESSAMENTO: BANCO DO BRASIL ===")
        logging.info(f"Arquivo Recebido: {arquivo}")
        def formatar_valor_brasileiro(valor):
            try:
                return locale.format_string("%.2f", float(valor), grouping=True)
            except:
                return valor
        try:
            extensao = arquivo.lower().split('.')[-1]
            logging.info(f"Extensão detectada: {extensao}")
            dados_importados = []
            saldo_final_calculado = 0

            if extensao == 'xls': #! SE O ARQUIVO É XLS
                logging.info("\n=== PROCESSANDO ARQUIVO XLS ===")
                wb = xlrd.open_workbook(arquivo)
                sheet = wb.sheet_by_index(0)
                logging.info(f"Planilha aberta: {sheet.name}")
                logging.info(f"Dimensões? {sheet.nrows} linhas x {sheet.ncols} colunas")

                logging.info("\nBuscando saldo inicial...")
                saldo_inicial = sheet.cell_value(3, 8)
                logging.info(f"Valor bruto encontrado em F4: {saldo_inicial}")
                logging.info(f"Tipo do valor: {type(saldo_inicial)}")

                if isinstance(saldo_inicial, str):
                    logging.info("Convertendo saldo inicial de string para float...")
                    saldo_inicial = float(saldo_inicial.replace(".", "").replace(",", "."))
                saldo_inicial_frmt = locale.format_string("%.2f", saldo_inicial, grouping=True)
                logging.info(f"Saldo inicial formatado: R${saldo_inicial_frmt}")

                resposta = messagebox.askyesno("Confirmação de saldo", f"O saldo inicial é de R${saldo_inicial_frmt}?")

                if not resposta:
                    logging.info("Usuário não confirmou o saldo inicial, solicitando entrada manual.")
                    saldo_inicial_manual = simpledialog.askstring("Entrada de Saldo", "Insira o saldo inicial correto:")
                    if saldo_inicial_manual:
                        try:
                            saldo_inicial = float(saldo_inicial_manual.replace(".", "").replace(",", "."))
                            saldo_inicial_frmt = locale.format_string("%.2f", saldo_inicial, grouping=True)
                        except ValueError:
                            messagebox.showerror("Erro", "Valor de saldo inicial inválido.")
                            return
                    else:
                        messagebox.showinfo("Aviso", "Processo cancelado pelo usuário.")
                        return
                    
                logging.info("Atualizando campo de saldo inicial na interface...")
                self.saldo_inicial_entry.delete(0, tk.END)
                self.saldo_inicial_entry.insert(0, saldo_inicial_frmt)

                saldo_final_calculado = saldo_inicial

                logging.info("\n=== INICIANDO PROCESSAMENTO DAS LINHAS ===")
                logging.info(f"Total de linhas na planilha: {sheet.nrows}")

                for row in range(4, sheet.nrows):
                    try:
                        logging.info(f"\nProcessando linha {row+1}:")
                        data = sheet.cell_value(row, 0)
                        logging.info(f"Data encontrada: {data} (tipo: {type(data)})")

                        historico = sheet.cell_value(row, 7)
                        num_doc = sheet.cell_value(row, 5)
                        valor = sheet.cell_value(row, 8)

                        logging.info(f"Valores lidos:")
                        logging.info(f"  Histórico: {historico}, N° Doc: {num_doc}, Valor: {valor}")

                        if valor is None or str(valor).strip() == "":
                            logging.info("Valor vazio, pulando linha...")
                            continue

                        def converter_para_float(valor):
                            logging.info(f"Tratando valor: {valor} (tipo: {type(valor)})")
                            if valor is None or valor == "":
                                logging.info("Valor vazio, retornando 0.0")
                                return 0.0
                            if isinstance(valor, str):
                                logging.info("Convertendo string para float...")
                                valor = valor.replace(".", "").replace(",", ".")
                            try:
                                resultado = float(valor)
                                logging.info(f"Valor convertido: {resultado}")
                                return resultado
                            except ValueError as e:
                                logging.info(f"Erro ao converter valor: {e}")
                                return 0.0
                            
                        valor_total = converter_para_float(valor)
                        valor_formatado = formatar_valor_brasileiro(valor_total)
                        logging.info(f"Valor total calculado: {valor_formatado}")

                        if isinstance(data, float):
                            logging.info("Convertendo data de float para string...")
                            data = xlrd.xldate_as_datetime(data, wb.datemode).strftime('%d/%m/%Y')
                            logging.info(f"Data convertida: {data}")
                            
                        logging.info("Adicionando linha aos dados importados...")
                        dados_importados.append([
                            data, historico, num_doc, valor_formatado, "",
                            "", "", "", "", "", "", "", ""
                        ])

                        saldo_final_calculado += valor_total
                        logging.info(f"Novo saldo calculado: {saldo_final_calculado}")
                            
                    except Exception as e:
                        logging.info(f"ERRO ao processar linha {row+1}:")
                        logging.info(f"Detalhes do erro: {str(e)}")
                        traceback.print_exc()
                        continue

            elif extensao == "xlsx": #! SE O ARQUIVO É XLSX
                logging.info("\n=== PROCESSANDO ARQUIVO XLSX ===")
                wb = openpyxl.load_workbook(arquivo, data_only=True)
                sheet = wb.active
                logging.info(f"Planilha ativa: {sheet.title}")

                logging.info("\nBuscando saldo inicial...")
                saldo_inicial_celula = sheet['I4'].value
                logging.info(f"Valor bruto encontrado em I4: {saldo_inicial_celula}")

                if isinstance(saldo_inicial_celula, str):
                    saldo_inicial = float(saldo_inicial_celula.replace(".", "").replace(",", "."))
                else:
                    saldo_inicial = float(saldo_inicial_celula)

                saldo_inicial_frmt = locale.format_string("%.2f", saldo_inicial, grouping=True)
                logging.info(f"Saldo inicial formatado: R${saldo_inicial_frmt}")

                resposta = messagebox.askyesno("Confirmação de saldo", f"O saldo inicial é de R${saldo_inicial_frmt}?")

                if not resposta:
                    logging.info("Usuário não confirmou o saldo inicial, solicitando entrada manual.")
                    saldo_inicial_manual = simpledialog.askstring("Entrada de Saldo", "Insira o saldo inicial correto:")
                    if saldo_inicial_manual:
                        try:
                            saldo_inicial = float(saldo_inicial_manual.replace(".", "").replace(",", "."))
                            saldo_inicial_frmt = locale.format_string("%.2f", saldo_inicial, grouping=True)
                        except ValueError:
                            messagebox.showerror("Erro", "Valor de saldo inicial inválido.")
                            return
                    else:
                        messagebox.showinfo("Aviso", "Processo cancelado pelo usuário.")
                        return
                    
                self.saldo_inicial_entry.delete(0, tk.END)
                self.saldo_inicial_entry.insert(0, saldo_inicial_frmt)

                logging.info("\n=== INICIANDO PROCESSAMENTO DAS LINHAS ===")
                for row in range(5, sheet.max_row + 1):
                    try:
                        logging.info(f"\nProcessando linha {row}:")
                        data = sheet.cell(row=row, column=1).value
                        logging.info(f"Data encontrada: {data}")

                        historico = sheet.cell(row=row, column=8).value
                        num_doc = sheet.cell(row=row, column=6).value
                        valor = sheet.cell(row=row, column=9).value
                        tipo_operacao = sheet.cell(row=row, column=10).value

                        logging.info(f"Valores lidos:")
                        logging.info(f"  Histórico: {historico}, N° Doc: {num_doc}, Valor: {valor}, Tipo: {tipo_operacao}")

                        def converter_para_float(valor):
                            if valor is None or valor == "":
                                return 0.0
                            if isinstance(valor, str):
                                valor = valor.replace(".", "").replace(",", ".")
                            try:
                                return float(valor)
                            except ValueError:
                                return 0.0
                            
                        valor_total = converter_para_float(valor)

                        if tipo_operacao and str(tipo_operacao).strip().upper() == "D":
                            valor_total = -abs(valor_total)
                        valor_formatado = formatar_valor_brasileiro(valor_total)
                        logging.info(f"Valor total calculado: {valor_total}")

                        dados_importados.append([
                            data, historico, num_doc, valor_formatado, "",
                            "", "", "", "", "", "", "", ""
                        ])

                        saldo_final_calculado += valor_total
                        logging.info(f"Novo saldo calculado: {saldo_final_calculado}")
                            
                    except Exception as e:
                        logging.info(f"ERRO ao processar linha {row}:")
                        logging.info(f"Detalhes do erro: {str(e)}")
                        continue

            logging.info("\n=== ATUALIZANDO INTERFACE ===")
            logging.info("Formatando saldo final...")
            saldo_final_calculado_frmt = locale.format_string("%.2f", saldo_final_calculado, grouping=True)
            logging.info(f"Saldo final formatado: R${saldo_final_calculado_frmt}")

            logging.info("Atualizando campo de saldo final...")
            self.saldo_final_calculado_entry.delete(0, tk.END)
            self.saldo_final_calculado_entry.insert(0, saldo_final_calculado_frmt)

            logging.info("\nLimpando Treeview...")
            for i in self.tree.get_children():
                self.tree.delete(i)

            logging.info("Inserindo dados na Treeview...")
            logging.info(f"Total de registros a inserir: {len(dados_importados)}")

            for i, dados in enumerate(dados_importados):
                tag = 'linha_par' if i % 2 == 0 else 'linha_impar'

                dados[3] = formatar_valor_brasileiro(dados[3])

                self.tree.insert("", "end", values=dados, tags=(tag,))

            self.atualizar_total_linhas_importadas()

        except Exception as e:
            logging.info("\n=== ERRO FATAL ===")
            logging.info(f"Erro: {str(e)}")
            logging.info("stack trace:")
            traceback.print_exc()
            messagebox.showerror("Erro",
                "Erro ao processar o arquivo. Verifique se:\n\n" +
                "1. O arquivo está no formato correto\n" +
                "2. O arquivo não está em modo de exibição protegida\n" +
                "3. O arquivo está fechado no Excel\n\n" +
                f"Erro: {str(e)}")
            return

    def acao_inter(self, arquivo):
        logging.info("\n=== INÍCIO DO PROCESSAMENTO: INTER ===")
        logging.info(f"Arquivo Recebido: {arquivo}")
        def formatar_valor_brasileiro(valor):
            try:
                return locale.format_string("%.2f", float(valor), grouping=True)
            except:
                return valor

        try:
            extensao = arquivo.lower().split('.')[-1]
            logging.info(f"Extensão detectada: {extensao}")
            dados_importados = []
            transacoes = []
            saldo_final_calculado = 0

            if extensao == "pdf": #! SE O ARQUIVO É PDF
                logging.info("\n=== PROCESSANDO ARQUIVO PDF ===")

                try:
                    import pdfplumber
                    import re
                    from tkinter import messagebox

                    def corrigir_valor(valor_str):
                        valor_str = valor_str.replace("R$", "").replace(".", "").replace(",", ".").replace(" ", "")
                        return float(valor_str)

                    def converter_data_extenso_para_ddmm(data_extenso):
                        meses = {
                            "janeiro": "01", "fevereiro": "02", "março": "03", "abril": "04",
                            "maio": "05", "junho": "06", "julho": "07", "agosto": "08",
                            "setembro": "09", "outubro": "10", "novembro": "11", "dezembro": "12"
                        }
                        match = re.match(r"(\d{1,2}) de (\w+) de (\d{4})", data_extenso.lower())
                        if match:
                            dia, mes_nome, ano = match.groups()
                            mes = meses.get(mes_nome)
                            if mes:
                                return f"{int(dia):02d}/{mes}/{ano}"
                        return data_extenso

                    data_regex = re.compile(r'^\d{1,2} de [A-Za-zçÇ]+ de \d{4}')
                    linhas_processadas = []

                    with pdfplumber.open(arquivo) as pdf:
                        for pagina in pdf.pages:
                            texto = pagina.extract_text()
                            if texto:
                                linhas_processadas.extend([linha.strip() for linha in texto.split('\n') if linha.strip()])

                    i = 0
                    while i < len(linhas_processadas):
                        linha = linhas_processadas[i]

                        if data_regex.match(linha):
                            data_linha = linha
                            data_formatada = converter_data_extenso_para_ddmm(data_linha)

                            descricao = ""
                            valor = 0.0
                            credito = ""
                            debito = ""

                            if i + 1 < len(linhas_processadas):
                                proxima_linha = linhas_processadas[i + 1].strip()

                                if not data_regex.match(proxima_linha):
                                    descricao = re.sub(r'-?R\$ ?[\d\.,]+.*', '', proxima_linha).strip()

                                    valor_match = re.search(r'-?\s*R\$ ?[\d\.,]+', proxima_linha)
                                    if not valor_match:
                                        valor_match = re.search(r'-?\s*R\$ ?[\d\.,]+', linha)

                                    if valor_match:
                                        valor_str = valor_match.group(0)

                                        valor = corrigir_valor(valor_str)

                                        logging.info(f"Valor bruto extraído: '{valor_str}' -> Valor corrigido: {valor}")

                                        valor_str_formatado = locale.format_string("%.2f", valor, grouping=True)

                                        credito = valor_str_formatado if valor > 0 else ""
                                        debito = valor_str_formatado if valor < 0 else ""

                                    i += 1

                            transacao = [data_formatada, descricao, "", credito or debito or "", "", "", "", "", "", "", "", "", ""]
                            transacoes.append(transacao)
                            dados_importados.append(transacao)
                            saldo_final_calculado += valor

                        i += 1

                    for transacao in dados_importados:
                        self.tree.insert("", "end", values=transacao)
                    
                    try:
                        saldo_final_str = self.saldo_final_entry.get().replace(".", "").replace(",", ".")
                        saldo_final = float(saldo_final_str)
                        diferenca = saldo_final - saldo_final_calculado

                        diferenca_frmt = locale.format_string("%.2f", diferenca, grouping=True)
                        self.diferenca_entry.delete(0, tk.END)
                        self.diference_entry.insert(0, diferenca_frmt)
                    except:
                        pass

                except Exception as e:
                    logging.info(f"❌ Erro ao processar PDF: {e}")
                    traceback.print_exc()
                    messagebox.showerror("Erro", f"Erro ao processar o PDF:\n\n{str(e)}")
                return
            
            elif extensao == "xls": #! SE O ARQUIVO É XLS
                logging.info("\n=== PROCESSANDO ARQUIVO XLS ===")

            elif extensao == "xlsx": #! SE O ARQUIVO É XLSX
                logging.info("\n=== PROCESSANDO ARQUIVO XLSX ===")

            logging.info("\n=== ATUALIZANDO INTERFACE ===")
            logging.info("Formatando saldo final...")
            saldo_final_calculado_frmt = locale.format_string("%.2f", saldo_final_calculado, grouping=True)
            logging.info(f"Saldo final formatado: R${saldo_final_calculado_frmt}")

            logging.info("Atualizando campo de saldo final...")
            self.saldo_final_calculado_entry.delete(0, tk.END)
            self.saldo_final_calculado_entry.insert(0, saldo_final_calculado_frmt)

            logging.info("\nLimpando Treeview...")
            for i in self.tree.get_children():
                self.tree.delete(i)

            logging.info("Inserindo dados na Treeview...")
            logging.info(f"Total de registros a inserir: {len(dados_importados)}")

            for i, dados in enumerate(dados_importados):
                tag = 'linha_par' if i % 2 == 0 else 'linha_impar'

                dados[3] = formatar_valor_brasileiro(dados[3])

                self.tree.insert("", "end", values=dados, tags=(tag,))

            self.atualizar_total_linhas_importadas()
                    
            logging.info("\n=== PROCESSAMENTO CONCLUÍDO COM SUCESSO ===")
            logging.info(f"Total de linhas processadas: {len(dados_importados)}")

        except Exception as e:
            logging.info("\n=== ERRO FATAL ===")
            logging.info(f"Erro: {str(e)}")
            logging.info("stack trace:")
            traceback.print_exc()
            messagebox.showerror("Erro",
                "Erro ao processar o arquivo. Verifique se:\n\n" +
                "1. O arquivo está no formato correto\n" +
                "2. O arquivo não está em modo de exibição protegida\n" +
                "3. O arquivo está fechado no Excel\n\n" +
                f"Erro: {str(e)}")
            return

    def acao_caixa(self, arquivo):
        logging.info('\n=== INÍCIO DO PROCESSAMENTO: CAIXA')
        logging.info(f"Arquivo recebido: {arquivo}")

        def formatar_valor_brasileiro(valor):
            try:
                return locale.format_string("%2.f", float(valor), grouping=True)
            except:
                return valor
            
        try:
            extensao = arquivo.lower().split('.')[-1]
            logging.info(f"Extensão detectada: {extensao}")
            dados_importados = []
            saldo_final_calculado = 0

            if extensao == 'pdf': #! SE O ARQUIVO É PDF
                logging.info("\n=== PROCESSANDO ARQUIVO PDF ===")
                try:
                    import pdfplumber
                    import re

                    with pdfplumber.open(arquivo) as pdf:
                        linhas = []
                        for pagina in pdf.pages:
                            texto = pagina.extract_text()
                            if texto:
                                linhas.extend(texto.split('\n'))

                    logging.info(f"📄 Total de linhas extraídas do PDF: {len(linhas)}")

                    data_regex = re.compile(r'^\d{2}/\d{2}/\d{4}')
                    transacoes = []

                    for i, linha in enumerate(linhas):
                        logging.info(f"🔹 [{i}] Lendo linha: {linha}")
                        if not isinstance(linha, str) or not data_regex.match(linha):
                            continue

                        if "SALDO ANTERIOR" in linha.upper():
                            logging.info("🔍 Encontrado 'SALDO ANTERIOR'. Iniciando captura nas próximas linhas...")
                            continue

                        if not data_regex.match(linha):
                            continue

                        match = re.match(r'^(\d{2}/\d{2}/\d{4})\s+(\d{6})\s+(.*?)\s+(\d+,\d{2})\s+([DC])\s+(\d+,\d{2})\s+([DC])$', linha)
                        if not match:
                            continue

                        data = match.group(1)
                        documento = match.group(2)
                        descricao = match.group(3).strip()
                        valor_str = match.group(4)
                        valor_tipo = match.group(5)
                        saldo_str = match.group(6)
                        saldo_tipo = match.group(7)

                        if descricao.upper() == "SALDO DIA":
                            continue

                        if match:
                            logging.info(match.groups())

                        valor_float = self.corrigir_valor(valor_str)
                        if valor_tipo == "D":
                            valor_float = -abs(valor_float)
                        else:
                            valor_float = abs(valor_float)

                        valor_str_formatado = locale.format_string("%.2f", abs(valor_float), grouping=True)

                        credito = valor_str_formatado if valor_float > 0 else ""
                        debito = valor_str_formatado if valor_float < 0 else ""

                        transacoes.append([
                            data, descricao, documento,
                            valor_str_formatado if valor_float > 0 else f"-{valor_str_formatado}",
                            saldo_str, "", "", "", "", "", "", "", "", ""
                        ])

                    for i, t in enumerate(transacoes):
                        tag = 'linha_par' if i % 2 == 0 else 'linha_impar'
                        while len(t) < 14:
                            t.append("")
                        self.tree.insert("", "end", values=t, tags=(tag,))

                    self.atualizar_total_linhas_importadas()
                    saldo_final_frmt = locale.format_string("%.2f", saldo_final_calculado, grouping=True)
                    self.saldo_final_calculado_entry.delete(0, tk.END)
                    self.saldo_final_calculado_entry.insert(0, saldo_final_frmt)

                    messagebox.showinfo("Sucesso", f"{len(transacoes)} transações importadas do PDF Caixa!")

                except Exception as e:
                    logging.info(f"❌ Erro ao processar PDF: {e}")
                    traceback.print_exc()
                    messagebox.showerror("Erro", f"Erro ao processar o PDF:\n\n{str(e)}")
                return

            elif extensao == 'xls': #! SE O ARQUIVO É XLS
                logging.info("\n=== PROCESSANDO ARQUIVO XLS ===")

            elif extensao == 'xlsx': #! SE O ARQUIVO É XLSX
                logging.info("\n=== PROCESSANDO ARQUIVO XLSX ===")

            logging.info("\n=== ATUALIZANDO INTERFACE ===")
            logging.info("Formatando saldo final...")
            saldo_final_calculado_frmt = locale.format_string("%.2f", saldo_final_calculado, grouping=True)
            logging.info(f"Saldo final formatado: R${saldo_final_calculado_frmt}")

            logging.info("Atualizando campo de saldo final...")
            self.saldo_final_calculado_entry.delete(0, tk.END)
            self.saldo_final_calculado_entry.insert(0, saldo_final_calculado_frmt)

            logging.info("\nLimpando Treeview...")
            for i in self.tree.get_children():
                self.tree.delete(i)

            logging.info("Inserindo dados na Treeview...")
            logging.info(f"Total de registros a inserir: {len(dados_importados)}")

            for i, dados in enumerate(dados_importados):
                tag = 'linha_par' if i % 2 == 0 else 'linha_impar'

                dados[3] = formatar_valor_brasileiro(dados[3])

                self.tree.insert("", "end", values=dados, tags=(tag,))

            self.atualizar_total_linhas_importadas()

            logging.info("\n=== PROCESSAMENTO CONCLUÍDO COM SUCESSO ===")
            logging.info(f"Total de linhas processadas: {len(dados_importados)}")

        except Exception as e:
            logging.info("\n=== ERRO FATAL ===")
            logging.info(f"Erro: {str(e)}")
            logging.info("Stack trace:")
            traceback.print_exc()
            messagebox.showerror("Erro",
                "Erro ao processar o arquivo. Verifique se:\n\n" +
                "1. O arquivo está no formato correto\n" +
                "2. O arquivo não está em modo de exibição protegida\n" +
                "3. O arquivo está fechado no Excel\n\n" +
                f"Erro: {str(e)}")
            return

    def acao_c6bank(self, arquivo):
        logging.info("\n=== INÍCIO DO PROCESSAMENTO: C6 BANK ===")
        logging.info(f"Arquivo Recebido: {arquivo}")
        def formatar_valor_brasileiro(valor):
            try:
                return locale.format_string("%.2f", float(valor), grouping=True)
            except:
                return valor
        try:
            extensao = arquivo.lower().split('.')[-1]
            logging.info(f"Extensão detectada: {extensao}")
            dados_importados = []
            saldo_final_calculado = 0

            if extensao == 'pdf': #! SE O ARQUIVO É PDF
                logging.info("\n=== PROCESSANDO ARQUIVO PDF ===")
                import pdfplumber
                import re

                with pdfplumber.open(arquivo) as pdf:
                    linhas = []
                    for pagina in pdf.pages:
                        texto = pagina.extract_text()
                        if texto:
                            linhas.extend(texto.split('\n'))

                logging.info(f"📄 Total de linhas extraídas do PDF: {len(linhas)}")

                regex_linha = re.compile(r'(\d{2}/\d{2})\s+.*?\s+(.+?)\s+(-?R?\$?\s*\d[\d\s\.,]*)$')

                transacoes = []

                for i, linha in enumerate(linhas):
                    logging.info(f"🔹 [{i}] Lendo linha: {linha}")

                    match = regex_linha.search(linha)
                    if not match:
                        continue

                    data = match.group(1)
                    descricao = match.group(2).strip()
                    valor_str = match.group(3).replace(" ", "").replace("R$", "").replace(".", "").replace(",", ".")
                
                    try:
                        valor = float(valor_str)
                    except ValueError:
                        logging.warning(f"⚠️ Erro ao converter valor: {valor_str}")
                        continue

                    credito = locale.currency(valor, grouping=True) if valor > 0 else ""
                    debito = locale.currency(-valor, grouping=True) if valor < 0 else ""

                    transacoes.append([
                        data, descricao, "",
                        credito,
                        debito,
                        "", "", "", "", "", "", "", "", ""
                    ])

                logging.info("\n=== ATUALIZANDO INTERFACE ===")
                self.tree.delete(*self.tree.get_children())

                for i, t in enumerate(transacoes):
                    tag = 'linha_par' if i % 2 == 0 else 'linha_impar'
                    while len(t) < 14:
                        t.append("")
                    self.tree.insert("", "end", values=t, tags=(tag,))

                self.atualizar_total_linhas_importadas()

                saldo_final = sum(float(t[3].replace("R$", "").replace(".", "").replace(",", ".")) if t[3] else -float(t[4].replace("R$", "").replace(".", "").replace(",", ".")) for t in transacoes)
                saldo_final_frmt = locale.currency(saldo_final, grouping=True)

                self.saldo_final_calculado_entry.delete(0, tk.END)
                self.saldo_final_calculado_entry.insert(0, saldo_final_frmt)

                logging.info("\n=== PROCESSAMENTO CONCLUÍDO COM SUCESSO ===")
                logging.info(f"Total de linhas processadas: {len(transacoes)}")
                messagebox.showinfo("Sucesso", f"{len(transacoes)} transações importadas com sucesso.")

            elif extensao == 'xls': #! SE O ARQUIVO É XLS
                logging.debug("\n=== PROCESSANDO ARQUIVO XLS ===")

            elif extensao == 'xlsx': #! SE O ARQUIVO É XLSX
                logging.debug("\n=== PROCESSANDO ARQUIVO XLSX ===")
            
            logging.info("\n=== ATUALIZANDO INTERFACE ===")
            logging.debug("Formatando saldo final...")
            saldo_final_calculado_frmt = locale.format_string("%.2f", saldo_final_calculado, grouping=True)
            logging.debug(f"Saldo final formatado: R${saldo_final_calculado_frmt}")

            logging.debug("Atualizando campo de saldo final...")
            self.saldo_final_calculado_entry.delete(0, tk.END)
            self.saldo_final_calculado_entry.insert(0, saldo_final_calculado_frmt)

            logging.debug("\nLimpando Treeview...")
            for i in self.tree.get_children():
                self.tree.delete(i)

            logging.debug("Inserindo dados na Treeview...")
            logging.debug(f"Total de registros a inserir: {len(dados_importados)}")

            for i, dados in enumerate(dados_importados):
                tag = 'linha_par' if i % 2 == 0 else 'linha_impar'

                dados[3] = formatar_valor_brasileiro(dados[3])

                self.tree.insert("", "end", values=dados, tags=(tag,))

            self.atualizar_total_linhas_importadas()
                    
            logging.info("\n=== PROCESSAMENTO CONCLUÍDO COM SUCESSO ===")
            logging.debug(f"Total de linhas processadas: {len(dados_importados)}")

        except Exception as e:
            logging.critical("\n=== ERRO FATAL ===")
            logging.critical(f"Erro: {str(e)}")
            logging.critical("stack trace:")
            traceback.print_exc()
            messagebox.showerror("Erro",
                "Erro ao processar o arquivo. Verifique se:\n\n" +
                "1. O arquivo está no formato correto\n" +
                "2. O arquivo não está em modo de exibição protegida\n" +
                "3. O arquivo está fechado no Excel\n\n" +
                f"Erro: {str(e)}")
            return

    #! ========== BANCOS QUE ESTÃO FALTANDO ==========

    def acao_grafeno(self, arquivo):
        logging.info("Executando ação específica para o Grafeno.")

    def acao_pagseguro(self, arquivo):
        logging.info("Executando ação específica para o Pagseguro.")

    def acao_hsbc(self, arquivo):
        logging.info("Executando ação específica para o HSBC.")

    def acao_suisse(self, arquivo):
        logging.info("Executando ação específica para o Credit Suisse.")

    def acao_sicredi(self, arquivo):
        logging.info("Executando ação específica para o Sicredi.")
    
    def editar_celula_treeview(self, event):
        item_id = self.tree.focus()
        if not item_id:
            return

        col = self.tree.identify_column(event.x)
        row = self.tree.identify_row(event.y)

        col_idx = int(col.replace('#', '')) - 1

        valores = list(self.tree.item(item_id, 'values'))

        x, y, width, height = self.tree.bbox(item_id, col)

        self.entry_edit = tk.Entry(self.frame_tree)
        self.entry_edit.place(x=x, y=y, width=width, height=height)
        self.entry_edit.insert(0, valores[col_idx])
        self.entry_edit.focus()

        def salvar_novo_valor(event):
            novo_valor = self.entry_edit.get()
            valores[col_idx] = novo_valor
            self.tree.item(item_id, values=valores)
            self.entry_edit.destroy()

        self.entry_edit.bind("<Return>", salvar_novo_valor)

        self.entry_edit.bind("<FocusOut>", lambda e: self.entry_edit.destroy())

    def confirmar_limpar_dados(self):
        resposta = messagebox.askyesno("Atenção", "Tem certeza que deseja limpar todos os dados?")
        if resposta:
            self.limpar_dados()

    def limpar_dados(self):
        self.saldo_inicial_entry.delete(0, tk.END)
        self.saldo_final_entry.delete(0, tk.END)
        self.saldo_final_calculado_entry.delete(0, tk.END)
        self.empresa_entry.delete(0, tk.END)
        self.conta_entry.delete(0, tk.END)
        self.diferenca_entry.delete(0, tk.END)
        self.banco_entry.delete(0, tk.END)
        self.agencia_conta_entry.delete(0, tk.END)
        self.saldo_final_contabil_entry.delete(0, tk.END)
        self.diferenca_extrato_bancario_entry.delete(0, tk.END)
        self.busca_entry.delete(0, tk.END)
        for i in self.tree.get_children():
            self.tree.delete(i)
        self.atualizar_total_linhas_importadas()

    def atualizar_total_linhas_importadas(self):
        total = len(self.tree.get_children())
        self.total_linhas_label.config(text=f"Linhas importadas: {total}")

        self.total_linhas_label.config(bg="#313131")

        self.root.after(300, lambda: self.total_linhas_label.config(bg="#313131"))

    def exportar_dados(self):
        wb = Workbook()
        ws = wb.active

        #* -------------------- DEFINE OS CABEÇALHOS DAS COLUNAS -------------------- #
        colunas_exportar = ["LancamentoLC", "DataLC", "DebitoLC", "D-C/CLC", "CreditoLC",
                      "C-C/CLC", "CNPJLC", "HistoricoLC", "ValorLC"]
        
        headers = {
            "LancamentoLC": "Lançamento",
            "DataLC": "Data",
            "DebitoLC": "Débito",
            "D-C/CLC": "D-C/C",
            "CreditoLC": "Crédito",
            "C-C/CLC": "C-C/C",
            "CNPJLC": "CNPJ",
            "HistoricoLC": "Histórico",
            "ValorLC": "Valor"
        }

        #* -------------------- ITERA SOBRE OS ITENS DA TREEVIEW E ADICIONA À PLANILHA -------------------- #
        dados = []
        for item in self.tree.get_children():
            values = self.tree.item(item, 'values')
            dados.append([values[self.tree["columns"].index(col)] for col in colunas_exportar])

        df = pd.DataFrame(dados, columns=[headers[col] for col in colunas_exportar])

        #* ------------------- SALVA O ARQUIVO EXCEL -------------------- #
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Arquivos Excel", "*.xlsx")])
        if file_path:
            df.to_excel(file_path, index=False)
            messagebox.showinfo("Exportar", "Dados exportados com sucesso!")

    def limpar_1p(self):
        resposta = messagebox.askyesno("Atenção", "Tem certeza que deseja limpar os dados da classificação?")
        if resposta:
            self.limpar_classificacao()

    def limpar_classificacao(self):
        colunas_para_limpar = [
            "LancamentoLC", "DataLC", "DebitoLC", "CreditoLC",
            "HistoricoLC", "ValorLC"
        ]
        
        for item in self.tree.get_children():
            values = list(self.tree.item(item, 'values'))
            for coluna in colunas_para_limpar:
                idx = self.tree["columns"].index(coluna)
                values[idx] = ""
            self.tree.item(item, values=values)

    def fechar_janela(self):
        if hasattr(self, 'tela_selecao_conta') and self.tela_selecao_conta.winfo_exists():
            self.tela_selecao_conta.root.destroy()
        self.root.destroy()

class TelaSelecaoConta:
    def __init__(self, root, callback):
        logging.info("Janela de seleção de conta iniciada.")
        self.root = root
        self.callback = callback
        self.root.title("Importador")
        self.root.geometry("480x210")
        self.root.config(bg='#313131')
        self.root.iconbitmap(r"K:\013 - Integracao\Macro\Importador de extratos\icon.ico")

        #* -------------------- TÍTULO PRINCIPAL -------------------- #
        self.title_label = tk.Label(root, text="Qual conta bancária irá importar?", font=("Roboto", 17, "bold"), bg='#313131') 
        self.title_label.grid(row=1, column=0, columnspan=1, padx=10, pady=10, sticky="w")

        #* -------------------- CAMPOS DE INFORMAÇÕES -------------------- #
        self.label_empresa = tk.Label(root, text="Empresa:", font=("Roboto", 10), bg='#313131') 
        self.label_empresa.grid(row=2, column=0, pady=1, padx=10, sticky="w")

        self.empresas = self.carregar_empresas()
        self.empresa_var = tk.StringVar()
        self.combobox_empresa = ttk.Combobox(root, textvariable=self.empresa_var, font=("Roboto", 10), width=61)
        self.combobox_empresa['values'] = self.empresas
        self.combobox_empresa.grid(row=3, column=0, pady=1, padx=10, sticky="w")
        self.combobox_empresa.bind("<KeyRelease>", self.filtrar_empresas)

        self.label_conta_contabil = tk.Label(root, text="Conta Contábil:", font=("Roboto", 10), bg='#313131')
        self.label_conta_contabil.grid(row=4, column=0, pady=1, padx=10, sticky="w")
        self.combobox_conta_contabil = ttk.Combobox(root, font=("Roboto", 10), width=61)
        self.combobox_conta_contabil.grid(row=5, column=0, pady=1, padx=10, sticky="w")
        self.combobox_conta_contabil.bind("<KeyRelease>", self.filtrar_contas)
        self.combobox_conta_contabil.bind("<FocusIn>", self.carregar_contas_ao_focar)

        #* -------------------- BOTÕES PRINCIPAIS -------------------- #
        self.btn_nova_empresa = ttk.Button(root, text="Nova Emp", command=self.abrir_tela_nova_empresa, width=9)
        self.btn_nova_empresa.grid(row=6, column=0, pady=5, padx=10, sticky="w")

        self.btn_nova_conta = ttk.Button(root, text="Nova Conta", command=self.abrir_tela_nova_conta, width=10)
        self.btn_nova_conta.grid(row=6, column=0, pady=5, padx=100, sticky="w")

        self.btn_alterar_conta = ttk.Button(root, text="Alterar Conta", command=self.alterar_conta, width=11)
        self.btn_alterar_conta.grid(row=6, column=0, pady=5, padx=197, sticky="w")

        self.btn_ok = ttk.Button(root, text="OK", command=self.confirmar, width=3)
        self.btn_ok.grid(row=6, column=0, pady=5, padx=105, sticky="e")

        self.btn_cancelar = ttk.Button(root, text="Cancelar", command=self.cancelar, width=8)
        self.btn_cancelar.grid(row=6, column=0, pady=5, padx=22, sticky="e")

    def carregar_contas_ao_focar(self, event=None):
        empresa = self.combobox_empresa.get()
        if empresa:
            self.atualizar_contas_contabeis()

    def atualizar_empresas(self):
        self.empresas = self.carregar_empresas()
        self.combobox_empresa['values'] = self.empresas
        self.combobox_empresa.set('')

    def filtrar_empresas(self, event=None):
        texto_digitado = self.combobox_empresa.get().lower()
        empresas_filtradas = [e for e in self.empresas if texto_digitado in e.lower()]
        self.combobox_empresa['values'] = empresas_filtradas

        if len(empresas_filtradas) == 1:
            self.combobox_empresa.set(empresas_filtradas[0])
            self.atualizar_contas_contabeis()

        if empresas_filtradas and not self.combobox_empresa.winfo_ismapped():
            self.root.after(100, lambda: self.combobox_empresa.event_generate('<Down>'))

    def filtrar_contas(self, event=None):
        texto_digitado = self.combobox_conta_contabil.get().lower()
        contas_disponiveis = self.combobox_conta_contabil['values']
        contas_filtradas = [c for c in contas_disponiveis if texto_digitado in c.lower()]

        self.combobox_conta_contabil['values'] =  contas_filtradas

        if len(contas_filtradas) == 1:
            self.combobox_conta_contabil.set(contas_filtradas[0])

        if contas_filtradas and not self.combobox_conta_contabil.winfo_ismapped():
            self.root.after(110, lambda: self.combobox_conta_contabil.event_generate('<Down>'))

        self.combobox_conta_contabil.bind("<KeyRelease>", self.filtrar_contas)

    def atualizar_contas_contabeis(self, event=None):
        empresa = self.combobox_empresa.get()
        if empresa:
            codigo_empresa = empresa.split(" - ")[0]
            contas = self.carregar_contas_contabeis(codigo_empresa)
            if contas:
                self.combobox_conta_contabil['values'] = contas
            else:
                self.combobox_conta_contabil['values'] = []
                resposta = messagebox.askyesno("Cadastrar Conta", "Nenhuma conta encontrada para esta empresa. Deseja cadastrar uma nova conta?")
                if resposta:
                    self.abrir_tela_nova_conta()
                else:
                    self.root.lift() #? deixa a janela importador como foco dps
        else:
            self.combobox_conta_contabil['values'] = []

    def carregar_empresas(self):
        caminho = caminho_dos_csv('empresas.csv')
        empresas = []
        if os.path.exists(caminho):
            with open(caminho, mode='r', newline='') as file:
                reader = csv.reader(file)
                for row in reader:
                    if len(row) >= 2:
                        empresas.append(f"{row[0]} - {row[1]}")
        logging.debug("Empresas carregadas:", empresas)
        return empresas

    def carregar_contas_contabeis(self, codigo_empresa):
        caminho = caminho_dos_csv('contas.csv')
        contas = []
        if os.path.exists(caminho):
            with open(caminho, mode='r', newline='') as file:
                reader = csv.reader(file)
                for row in reader:
                    if len(row) >= 6 and row[5] == codigo_empresa:
                        descricao = f"{row[0]}, {row[1]}, {row[2]}, {row[3]}, {row[5]}"
                        contas.append(descricao)
        logging.info("Empresa selecionada: {codigo_empresa}")
        logging.debug("Contas encontradas: ", contas)
        return contas

    def abrir_tela_nova_empresa(self):
        if not hasattr(self, 'tela_nova_empresa') or not self.tela_nova_empresa.winfo_exists():
            self.tela_nova_empresa = tk.Toplevel(self.root)
            app = TelaNovaEmpresa(self.tela_nova_empresa, self.salvar_nova_empresa)

    def abrir_tela_nova_conta(self):
        if not hasattr(self, 'tela_nova_conta') or not self.tela_nova_conta.winfo_exists():
            self.tela_nova_conta = tk.Toplevel(self.root)
            app = TelaNovaConta(self.tela_nova_conta, self.salvar_nova_conta)

    def salvar_nova_empresa(self, codigo, razao_social):
        caminho = caminho_dos_csv('empresas.csv')
        with open(caminho, mode='a', newline='') as file:
            writer = csv.writer(file)
            writer.writerow([codigo, razao_social])
        messagebox.showinfo("Nova Empresa", "Empresa adicionada com sucesso!")
        self.combobox_empresa['values'] = self.carregar_empresas()


    def salvar_nova_conta(self, codigo_empresa, banco, agencia, conta_bancaria, conta_ativo, conta_passivo):
        caminho = caminho_dos_csv('contas.csv')
        with open(caminho, mode='a', newline='') as file:
            writer = csv.writer(file)
            writer.writerow([codigo_empresa, banco, agencia, conta_bancaria, conta_ativo, conta_passivo])
        messagebox.showinfo("Nova Conta", "Conta adicionada com sucesso!")
        self.combobox_conta_contabil['values'] = self.carregar_contas_contabeis()


    def verificar_empresa(self, event):
        empresa = self.combobox_empresa.get()
        if empresa:
            codigo = empresa.split(" - ")[0]
            if not any(codigo in e for e in self.empresas):
                self.perguntar_cadastrar_empresa(codigo)
            else:
                self.combobox_conta_contabil.focus_set()

    def verificar_empresa_digitada(self, event):
        empresa = self.combobox_empresa.get()
        if empresa:
            codigo = empresa.split(" - ")[0]
            if not any(codigo in e for e in self.empresas):
                self.perguntar_cadastrar_empresa(codigo)

    def perguntar_cadastrar_empresa(self, codigo):
        resposta = messagebox.askyesno("Cadastrar Empresa", f"A empresa com código {codigo} não está cadastrada. Deseja cadastrá-la?")
        if resposta:
            self.abrir_tela_nova_empresa()

    def confirmar(self):
        empresa = self.combobox_empresa.get()
        conta = self.combobox_conta_contabil.get()
        if empresa and conta:
            arquivo = None
            respostas_safra = {}  #? dicionário para armazenar as respostas
            
            #? verificar se é conta Safra
            if "Safra" in conta:
                resp1 = messagebox.askyesno("Banco Safra", "O extrato está no formato PDF?")
                if resp1:
                    logging.info("O usuário escolheu a opção PDF.")
                    resp11 = messagebox.askyesno("Banco Safra", "O extrato está no novo formato do Banco Safra?")
                    if resp11:
                        resp12 = messagebox.askyesno("Banco Safra", "O extrato é conta vinculada?")
                        arquivo = filedialog.askopenfilename(filetypes=[("Arquivos PDF", "*.pdf")])
                        #? armazenar as respostas para usar na acao_safra
                        respostas_safra = {
                            "novo_formato": True,
                            "conta_vinculada": resp12
                        }
                    else:
                        arquivo = filedialog.askopenfilename(filetypes=[("Arquivos PDF", "*.pdf")])
                        respostas_safra = {
                            "novo_formato": True,
                            "conta_vinculada": False
                        }




                else:
                    resp2 = messagebox.askyesno("Banco Safra", "O extrato está no novo formato do Banco Safra?")
                    if resp2:
                        resp3 = messagebox.askyesno("Banco Safra", "O extrato é conta vinculada?")
                        arquivo = filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xls;*.xlsx")])
                        #? armazenar as respostas para usar na acao_safra
                        respostas_safra = {
                            "novo_formato": True,
                            "conta_vinculada": resp3
                        }
                    else:
                        arquivo = filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xls;*.xlsx")])
                        respostas_safra = {
                            "novo_formato": False,
                            "conta_vinculada": False
                        }
            else:
                pdf_resposta = messagebox.askyesno("Formato do Extrato", "O extrato está em PDF?")
                if pdf_resposta:
                    arquivo = filedialog.askopenfilename(filetypes=[("Arquivos PDF", "*.pdf")])
                else:
                    xlsx_resposta = messagebox.askyesno("Formato do Extrato", "O extrato está em XLS ou XLSX?")
                    if xlsx_resposta:
                        arquivo = filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xls;*.xlsx")])
                    else:
                        messagebox.showinfo("Aviso", "Por favor, selecione um arquivo válido.")

            if arquivo:
                #? passar as respostas junto com os outros parâmetros
                self.callback(empresa, conta, arquivo, respostas_safra)

            self.root.destroy()
        else:
            messagebox.showwarning("Aviso", "Por favor, preencha ambos os campos.")

    def cancelar(self):
        self.root.destroy()

    def alterar_conta(self):
        messagebox.showinfo("Alterar Conta", "Funcionalidade de Alterar Conta ainda não implementada.")
        self.root.lift()
        
class TelaNovaConta:
    def __init__(self, root, callback):
        logging.info("Janela para criação de conta iniciada.")
        self.root = root
        self.callback = callback
        self.root.title("Nova Conta")
        self.root.geometry("460x215")
        self.root.config(bg='#313131')
        self.root.iconbitmap(r"K:\013 - Integracao\Macro\Importador de extratos\icon.ico")

        #* -------------------- TÍTULO PRINCIPAL -------------------- #
        self.title_label = tk.Label(root, text="Informações da conta bancária", font=("Roboto", 17, "bold"), bg='#313131') 
        self.title_label.grid(row=1, column=0, columnspan=1, padx=10, pady=10, sticky="w")

        #* -------------------- CAMPOS DE INFORMAÇÕES -------------------- #
        self.label_codigo_empresa = tk.Label(root, text="Código da Empresa:", font=("Roboto", 10), bg='#313131') 
        self.label_codigo_empresa.grid(row=2, column=0, pady=1, padx=10, sticky="w")
        self.entry_codigo_empresa = ttk.Entry(root, font=("Roboto", 10), width=15)
        self.entry_codigo_empresa.grid(row=3, column=0, pady=1, padx=10, sticky="w")

        self.label_bancos = tk.Label(root, text="Banco:", font=("Roboto", 10), bg='#313131')
        self.label_bancos.grid(row=2, column=0, pady=1, padx=134, sticky="w")
        self.bancos = ["001 - Banco do Brasil", "077 - Banco Inter", "104 - Banco Caixa Eletronica", 
                       "237 - Banco Bradesco", "274 - Banco Grafeno", "290 - Banco Pagseguro", 
                       "336 - Banco C6 Bank", "341 - Banco Itau", "353 - Banco Santander", 
                       "399 - Banco HSBC", "422 - Banco Safra", "505 - Credit Suisse", 
                       "707 - Banco Daycoval", "748 - Banco Sicredi"]
        self.combobox_banco = ttk.Combobox(root, values=self.bancos, font=("Roboto", 10), width=40)
        self.combobox_banco.grid(row=3, column=0, pady=1, padx=135, sticky="w")

        self.label_agencia = tk.Label(root, text="Agência:", font=("Roboto", 10), bg='#313131')
        self.label_agencia.grid(row=6, column=0, pady=1, padx=10, sticky="w")
        self.entry_agencia = ttk.Entry(root, font=("Roboto", 10), width=10)
        self.entry_agencia.grid(row=7, column=0, pady=1, padx=10, sticky="w")

        self.label_conta_bancaria = tk.Label(root, text="Conta Bancária:", font=("Roboto", 10), bg='#313131')
        self.label_conta_bancaria.grid(row=6, column=0, pady=1, padx=100, sticky="w")
        self.entry_conta_bancaria = ttk.Entry(root, font=("Roboto", 10), width=15)
        self.entry_conta_bancaria.grid(row=7, column=0, pady=1, padx=100, sticky="w")

        self.label_conta_ativo = tk.Label(root, text="N° Conta Ativo:", font=("Roboto", 10), bg='#313131')
        self.label_conta_ativo.grid(row=6, column=0, pady=1, padx=225, sticky="w")
        self.entry_conta_ativo = ttk.Entry(root, font=("Roboto", 10), width=13)
        self.entry_conta_ativo.grid(row=7, column=0, pady=1, padx=225, sticky="w")

        self.label_conta_passivo = tk.Label(root, text="N° Conta Passivo:", font=("Roboto", 10), bg='#313131')
        self.label_conta_passivo.grid(row=6, column=0, pady=1, padx=145, sticky="e")
        self.entry_conta_passivo = ttk.Entry(root, font=("Roboto", 10), width=14)
        self.entry_conta_passivo.grid(row=7, column=0, pady=1, padx=135, sticky="e")

        #* -------------------- BOTÕES PRINCIPAIS -------------------- #
        self.btn_salvar = ttk.Button(root, text="Salvar", command=self.salvar, width=8)
        self.btn_salvar.grid(row=8, column=0, sticky="e", padx=220, pady=8)

        self.btn_cancelar = ttk.Button(root, text="Cancelar", command=self.cancelar, width=8)
        self.btn_cancelar.grid(row=8, column=0, sticky="e", padx=135, pady=8)

    def salvar(self):
        codigo_empresa = self.entry_codigo_empresa.get()
        bancos = self.combobox_banco.get()
        agencia = self.entry_agencia.get()
        conta_bancaria = self.entry_conta_bancaria.get()
        conta_ativo = self.entry_conta_ativo.get()
        conta_passivo = self.entry_conta_passivo.get()
        if codigo_empresa and bancos and agencia and conta_bancaria and conta_ativo and conta_passivo:
            self.root.destroy()
            self.callback(conta_ativo, bancos, agencia, conta_bancaria, conta_passivo, codigo_empresa)
        else:
            messagebox.showwarning("Aviso", "Por favor, preencha todos os campos.")

    def cancelar(self):
        self.root.destroy()

class TelaNovaEmpresa:
    def __init__(self, root, callback):
        logging.info("Janela para criação de empresa iniciada.")
        self.root = root
        self.callback = callback
        self.root.title("Nova Empresa")
        self.root.geometry("480x167")
        self.root.config(bg='#313131')
        self.root.iconbitmap(r"K:\013 - Integracao\Macro\Importador de extratos\icon.ico")

        #* -------------------- TÍTULO PRINCIPAL -------------------- #
        self.title_label = tk.Label(root, text="Informe os dados da empresa", font=("Roboto", 17, "bold"), bg='#313131') 
        self.title_label.grid(row=1, column=0, columnspan=1, padx=10, pady=10, sticky="w")

        #* -------------------- CAMPOS DE INFORMAÇÕES -------------------- #
        self.label_codigo = tk.Label(root, text="Código:", font=("Roboto", 10), bg='#313131') 
        self.label_codigo.grid(row=2, column=0, pady=1, padx=10, sticky="w")
        self.entry_codigo = ttk.Entry(root, font=("Roboto", 10), width=6) 
        self.entry_codigo.grid(row=3, column=0, pady=1, padx=10, sticky="w")

        self.label_razao_social = tk.Label(root, text="Razão Social:", font=("Roboto", 10), bg='#313131') 
        self.label_razao_social.grid(row=2, column=0, pady=1, padx=80, sticky="w")
        self.entry_razao_social = ttk.Entry(root, font=("Roboto", 10), width=53) 
        self.entry_razao_social.grid(row=3, column=0, pady=1, padx=80, sticky="w")

        #* -------------------- BOTÕES PRINCIPAIS -------------------- #
        self.btn_salvar = ttk.Button(root, text="Salvar", command=self.salvar, width=8)
        self.btn_salvar.grid(row=7, column=0, columnspan=2, pady=10, padx=164, sticky="e")

        self.btn_cancelar = ttk.Button(root, text="Cancelar", command=self.cancelar, width=8)
        self.btn_cancelar.grid(row=7, column=0, columnspan=2, pady=10, padx=80, sticky="e")

    def salvar(self):
        codigo = self.entry_codigo.get()
        razao_social = self.entry_razao_social.get()
        if codigo and razao_social:
            self.root.destroy()
            self.callback(codigo, razao_social)
        else:
            messagebox.showwarning("Aviso", "Por favor, preencha ambos os campos.")

    def cancelar(self):
        self.root.destroy()

if __name__ == "__main__":
    login_root = tk.Tk()
    login_app = SenhaLogin(login_root, open_main_window)
    login_root.mainloop()